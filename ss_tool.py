import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go
import plotly.express as px
import plotly.io as pio

st.set_page_config(page_title="Sale Analyzer", layout="wide", page_icon="📊")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Plus+Jakarta+Sans:wght@600;700;800&display=swap');

*, *::before, *::after { font-family: 'Inter', sans-serif !important; box-sizing: border-box; }
.stApp { background: #f4f0ff !important; }
#MainMenu, footer, header { visibility: hidden; }

/* Remove Streamlit default top padding */
.block-container { padding-top: 0.8rem !important; padding-bottom: 1rem !important; }
[data-testid="stAppViewContainer"] > section > div { padding-top: 0 !important; }

/* ══ NAVBAR ══ */
.hero {
    padding: 0.55rem 1.4rem;
    display: flex; align-items: center; gap: 1rem;
    background: linear-gradient(90deg, #3a0068 0%, #6a1b9a 55%, #9c27b0 100%);
    margin-bottom: 1rem; border-radius: 12px;
    box-shadow: 0 3px 14px rgba(106,27,154,0.3);
}
.hero-badge {
    background: rgba(255,255,255,0.18); border: 1.5px solid rgba(255,255,255,0.35);
    color: #ffffff; font-size:.56rem; font-weight:700; letter-spacing:2px;
    text-transform:uppercase; padding:4px 11px; border-radius:20px;
    white-space:nowrap; flex-shrink:0;
}
.hero-divider {
    width:1px; height:26px; background:rgba(255,255,255,0.22); flex-shrink:0;
}
.hero-title {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-size: 1.05rem; font-weight: 800; color: #ffffff;
    margin: 0; line-height: 1; white-space:nowrap; flex-shrink:0;
}
.hero-arrow { color:rgba(255,255,255,0.45); font-size:.95rem; flex-shrink:0; }
.hero-sub-line {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-size: .8rem; font-weight: 600; color: #e8c8ff;
    white-space:nowrap; overflow:hidden; text-overflow:ellipsis;
}
.hero-sub {
    color: rgba(255,255,255,0.52); font-size: .65rem;
    margin: 0.08rem 0 0 0; font-weight: 400;
}

/* ══ KPI CARDS ══ */
.kpi-card {
    background: linear-gradient(135deg, #6a1b9a 0%, #9c27b0 100%);
    border: none; border-radius: 16px;
    padding: 1.2rem 1.4rem;
    box-shadow: 0 4px 18px rgba(106,27,154,0.35);
    transition: box-shadow 0.2s, transform 0.2s;
}
.kpi-card:hover { box-shadow: 0 8px 28px rgba(106,27,154,0.45); transform: translateY(-3px); }
.kpi-label { font-size:.6rem; font-weight:700; letter-spacing:2.5px; text-transform:uppercase; color:rgba(255,255,255,0.75); margin-bottom:.5rem; }
.kpi-value { font-family:'Plus Jakarta Sans', sans-serif !important; font-size:1.75rem; font-weight:800; color:#ffffff; line-height:1; }
.kpi-value.blue { color:#ffffff !important; }
.kpi-sub { font-size:.75rem; color:rgba(255,255,255,0.7); margin-top:.35rem; font-weight:500; }

/* ══ SECTION TITLES ══ */
.section-title {
    font-size:.65rem; font-weight:700; letter-spacing:2.5px; text-transform:uppercase;
    color:#6a1b9a; padding:.5rem 0; margin-bottom:.8rem;
    border-bottom: 2.5px solid #ddd6fe;
}

/* ══ STREAMLIT OVERRIDES ══ */
.stApp > div { background: #f4f0ff !important; }

/* Text - make ALL dark and readable */
p { color: #1a0030 !important; font-size:.9rem !important; }
label { color: #3d0066 !important; font-weight:600 !important; font-size:.85rem !important; }
[data-testid="stWidgetLabel"] p { color:#6a1b9a !important; font-size:.85rem !important; font-weight:600 !important; }
div[data-testid="stDataFrame"] * { color:#1a0030 !important; font-size:.84rem !important; }

/* Inputs */
.stSelectbox > div > div, .stMultiSelect > div > div {
    background:#ffffff !important; border:1.5px solid #c084fc !important; border-radius:10px !important;
    color:#1a0030 !important;
}
[data-baseweb="popover"] { background:#fff !important; border:1px solid #c084fc !important; }
[data-baseweb="popover"] * { color:#1a0030 !important; background:#fff !important; }
[data-baseweb="option"]:hover { background:rgba(106,27,154,0.08) !important; }
li[aria-selected="true"] { background:rgba(106,27,154,0.12) !important; color:#6a1b9a !important; }

/* Tags in multiselect */
[data-baseweb="tag"] { background:#ede9fe !important; }
[data-baseweb="tag"] span { color:#4c1d95 !important; font-weight:600 !important; }

/* Buttons */
.stButton > button {
    background: linear-gradient(135deg,#6a1b9a,#9c27b0) !important;
    color: #ffffff !important; border: none !important; border-radius: 12px !important;
    font-weight: 700 !important; font-size: .95rem !important;
    padding: .7rem 2rem !important;
    box-shadow: 0 4px 14px rgba(106,27,154,0.38) !important;
    transition: all 0.2s !important; letter-spacing: .4px !important;
    display: flex !important; align-items: center !important; justify-content: center !important;
    text-shadow: 0 1px 2px rgba(0,0,0,0.2) !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg,#7b1fa2,#ab47bc) !important;
    box-shadow: 0 6px 22px rgba(106,27,154,0.48) !important;
    transform: translateY(-1px) !important;
    color: #ffffff !important;
}
.stButton > button p { color: #ffffff !important; font-weight: 700 !important; }
.stDownloadButton > button {
    background:#fff !important; color:#6a1b9a !important;
    border:2px solid #6a1b9a !important; border-radius:10px !important;
    font-weight:700 !important; font-size:.88rem !important;
}
.stDownloadButton > button:hover { background:#f5f0ff !important; }

/* ══ FILE UPLOADER ══ */
[data-testid="stFileUploader"] {
    background: linear-gradient(90deg, #3a0068 0%, #6a1b9a 55%, #9c27b0 100%) !important;
    border:2px dashed rgba(255,255,255,0.4) !important; border-radius:14px !important;
}
[data-testid="stFileUploaderDropzone"] {
    background: linear-gradient(90deg, #3a0068 0%, #6a1b9a 55%, #9c27b0 100%) !important;
    border: none !important;
}
[data-testid="stFileUploader"] * { color:#ffffff !important; font-weight:500 !important; }
[data-testid="stFileUploaderFileName"] { color:#ffffff !important; font-weight:700 !important; }
[data-testid="stFileUploaderDropzone"] svg { fill: #ffffff !important; }
[data-testid="stFileUploaderDropzone"] button {
    visibility: visible !important;
    height: auto !important;
    padding: .35rem 1rem !important;
    margin: .4rem auto 0 !important;
    background: rgba(255,255,255,0.2) !important;
    border: 1.5px solid rgba(255,255,255,0.5) !important;
    border-radius: 8px !important;
    color: #ffffff !important;
    font-weight: 700 !important;
    font-size: .8rem !important;
    display: block !important;
}
[data-testid="stFileUploaderDropzone"] button:hover {
    background: rgba(255,255,255,0.35) !important;
}
[data-testid="stFileUploadDeleteBtn"] button,
[data-testid="stFileUploadDeleteBtn"] {
    visibility: visible !important;
    height: auto !important;
    padding: 0 !important;
    color: #ffffff !important;
    opacity: 0.85 !important;
}
[data-testid="stFileUploadDeleteBtn"] svg { fill: #ffffff !important; }

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    background:#fff !important; border-radius:12px !important; padding:4px !important;
    border:1.5px solid #ddd6fe !important; box-shadow: 0 2px 8px rgba(106,27,154,0.08) !important;
}
.stTabs [data-baseweb="tab"] {
    color: #1a0030 !important; border-radius:8px !important;
    font-size:.84rem !important; font-weight:600 !important;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg,#6a1b9a,#9c27b0) !important;
    color: #ffffff !important; font-weight:700 !important;
}
.stTabs [aria-selected="true"] * { color: #ffffff !important; }

/* Success/spinner */
.stSuccess { background:#f0fdf4 !important; border:1px solid #86efac !important; border-radius:10px !important; }
.stSuccess * { color:#166534 !important; font-weight:600 !important; }
[data-testid="stSpinner"] * { color:#6a1b9a !important; }

/* Dataframe */
[data-testid="stFileUploaderDropzone"] button { display: none !important; }
</style>
""", unsafe_allow_html=True)

# ══════════════════ CONSTANTS ══════════════════
MONTHS = ["1-April'25","2-May'25","3-June'25","4-July'25","5-Aug'25",
          "6-Sep'25","7-Oct'25","8-Nov'25","9-Dec'25","10-Jan'26","11-Feb'26"]
MONTH_SHORT = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb"]
CATS = ['BAG','BANDS','COLD WEATHER','HYDRATION','INFLATABLE',
        'MOVEMENT SUPPORT','OTHER','PERFORMANCE APPAREL','RUNNING','TRAINING','YOGA']

GOLD = "1565c0"; WHITE = "FFFFFF"; BG1 = "1e3a5f"; BG2 = "e8f0f9"
BG3 = "dce8f5"; BG4 = "c5d5e8"; GREY = "607d9b"

def chart_layout(height=400, title="", xangle=0, show_legend=True):
    """Returns a clean layout dict — no xaxis/yaxis conflict"""
    return dict(
        paper_bgcolor="rgba(255,255,255,1)",
        plot_bgcolor="rgba(245,240,255,0.5)",
        font=dict(color="#1a0030", family="Inter", size=12),
        margin=dict(l=10, r=10, t=55, b=10),
        height=height,
        title=dict(text=f"<b>{title}</b>", font=dict(color="#1a0030", size=15, family="Plus Jakarta Sans")),
        legend=dict(
            font=dict(color="#1a0030", size=11),
            bgcolor="rgba(255,255,255,0.97)",
            bordercolor="#ddd6fe", borderwidth=1.5,
            visible=show_legend,
        ),
        xaxis=dict(
            gridcolor="#ede9fe", tickfont=dict(color="#1a0030", size=11, family="Inter"),
            linecolor="#ddd6fe", tickangle=xangle, showgrid=True,
        ),
        yaxis=dict(
            gridcolor="#ede9fe", tickfont=dict(color="#1a0030", size=11, family="Inter"),
            linecolor="#ddd6fe", showgrid=True,
        ),
    )

# Chart color palettes
BLUE_SEQ = [[0,'#f3e5f5'],[0.4,'#9c27b0'],[1,'#6a1b9a']]
CAT_COLORS_LIGHT = ['#7b1fa2','#e91e63','#ff6f00','#1565c0','#2e7d32',
                    '#00838f','#f57f17','#6a1b9a','#c62828','#00695c','#4527a0']

# ══════════════════ HELPERS ══════════════════
def fmt_inr(v):
    if pd.isna(v) or v == 0: return "—"
    v = int(round(float(v)))
    s = str(abs(v)); prefix = "-" if v < 0 else ""
    if len(s) <= 3: return prefix + s
    last3 = s[-3:]; rest = s[:-3]; groups = []
    while len(rest) > 2: groups.append(rest[-2:]); rest = rest[:-2]
    if rest: groups.append(rest)
    groups.reverse()
    return prefix + ','.join(groups) + ',' + last3

def pct(v, dec=2):
    if pd.isna(v) or v == 0: return "—"
    return f"{float(v)*100:.{dec}f}%"

def fill(hex_c): return PatternFill("solid", fgColor=hex_c)
def font(color="1a2e4a", size=9, bold=False): return Font(bold=bold, size=size, color=color, name="Calibri")
def align(h="center", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)
def border():
    s = Side(style='thin', color="c5d5e8")
    return Border(left=s, right=s, top=s, bottom=s)
def thick_border():
    t = Side(style='medium', color="1565c0")
    s = Side(style='thin', color="c5d5e8")
    return Border(left=s, right=s, top=t, bottom=t)

# ── Excel Color Palette (Pastel Multi-color like CRM dashboard) ──
E_TITLE_BG   = "1e3a5f"   # Dark navy  - title row bg
E_TITLE_FG   = "FFFFFF"   # White      - title text
E_HDR_BG     = "f8f9fa"   # Light grey - header bg
E_HDR_FG     = "374151"   # Dark grey  - header text
E_SUBHDR_BG  = "e8f4fd"   # Light blue - sub-header bg
E_SUBHDR_FG  = "1565c0"   # Blue
E_CONT_BG    = "f8f9fa"   # Light grey - contribution row
E_CONT_FG    = "6b7280"   # Grey       - contribution text
E_TOTAL_BG   = "e8eaf6"   # Light indigo - grand total bg
E_TOTAL_FG   = "1a237e"   # Dark indigo - grand total text
E_ROW1       = "FFFFFF"   # White      - odd rows
E_ROW2       = "FFFFFF"   # White      - even rows (no stripe, use col colors)
E_DATA_FG    = "374151"   # Dark grey  - data text
E_SALE_FG    = "1565c0"   # Blue       - Total Sale highlight
E_STK_FG     = "6b7280"   # Grey       - Stock text
E_CONT_PCT   = "374151"   # Dark       - contribution %
E_BORDER     = "e5e7eb"   # Light grey border

# Pastel column colors (like the CRM example - each month gets a color band)
MONTH_COLORS = [
    "dbeafe",  # light blue    - Apr
    "dcfce7",  # light green   - May
    "fef9c3",  # light yellow  - Jun
    "ffedd5",  # light orange  - Jul
    "fce7f3",  # light pink    - Aug
    "ede9fe",  # light purple  - Sep
    "d1fae5",  # light teal    - Oct
    "fee2e2",  # light red     - Nov
    "dbeafe",  # light blue    - Dec
    "dcfce7",  # light green   - Jan
    "fef9c3",  # light yellow  - Feb
]
MONTH_FG = [
    "1e40af",  # blue    - Apr
    "166534",  # green   - May
    "854d0e",  # yellow  - Jun
    "9a3412",  # orange  - Jul
    "9d174d",  # pink    - Aug
    "5b21b6",  # purple  - Sep
    "065f46",  # teal    - Oct
    "991b1b",  # red     - Nov
    "1e40af",  # blue    - Dec
    "166534",  # green   - Jan
    "854d0e",  # yellow  - Feb
]
# CWC category pastel colors (Sale col)
CAT_COLORS_XL = [
    ("dbeafe","1e40af"),   # BAG - blue
    ("dcfce7","166534"),   # BANDS - green
    ("fef9c3","854d0e"),   # COLD WEATHER - yellow
    ("ffedd5","9a3412"),   # HYDRATION - orange
    ("fce7f3","9d174d"),   # INFLATABLE - pink
    ("ede9fe","5b21b6"),   # MOVEMENT SUPPORT - purple
    ("d1fae5","065f46"),   # OTHER - teal
    ("fee2e2","991b1b"),   # PERFORMANCE APPAREL - red
    ("dbeafe","1e40af"),   # RUNNING - blue
    ("dcfce7","166534"),   # TRAINING - green
    ("fef9c3","854d0e"),   # YOGA - yellow
]

# ══════════════════ PROCESSING ══════════════════
def process(file):
    df = pd.read_excel(file, header=1)
    df.columns = [str(c).strip() for c in df.columns]
    c = list(df.columns); c[-1] = 'Remarks2'; df.columns = c
    df['Mrp Vlu'] = pd.to_numeric(df['Mrp Vlu'], errors='coerce').fillna(0)

    sale  = df[df['Sale/ Stock'] == 'Sale'].copy()
    stock = df[df['Sale/ Stock'] == 'Stock'].copy()


    # ── SWC ──
    swc = sale.pivot_table(index='Store Name', columns='Month',
                           values='Mrp Vlu', aggfunc='sum').reindex(columns=MONTHS)
    swc['Total Sale'] = swc[MONTHS].sum(axis=1)
    closing = stock[stock['Month']=='Feb Closing'].groupby('Store Name')['Mrp Vlu'].sum()
    swc['Feb Closing Stk'] = closing
    grand = swc['Total Sale'].sum()
    swc['Sale Cont.'] = swc['Total Sale'] / grand

    # month cont row
    mc = (swc[MONTHS].sum() / grand).to_frame().T
    mc.index = ['_mc']
    mc['Total Sale'] = 1.0; mc['Feb Closing Stk'] = np.nan; mc['Sale Cont.'] = np.nan

    gt_swc = swc.sum().to_frame().T
    gt_swc.index = ['Grand Total']
    gt_swc['Sale Cont.'] = 1.0

    swc_final = pd.concat([mc, swc, gt_swc])

    # ── CWC ──
    avail = [c for c in CATS if c in sale['CATEGORY'].unique()]
    cwc_s = sale.pivot_table(index='Store Name', columns='CATEGORY',
                             values='Mrp Vlu', aggfunc='sum').reindex(columns=avail).fillna(0)
    cwc_k = stock[stock['Month']=='Feb Closing'].pivot_table(
                index='Store Name', columns='CATEGORY',
                values='Mrp Vlu', aggfunc='sum').reindex(columns=avail).fillna(0)
    cwc_s['TOTAL'] = cwc_s.sum(axis=1)
    cwc_k['TOTAL'] = cwc_k.sum(axis=1)

    gt_s = cwc_s.sum(); gt_s.name = 'Grand Total'
    gt_k = cwc_k.sum(); gt_k.name = 'Grand Total'

    cont_s = (gt_s / gt_s['TOTAL']); cont_s.name = 'Contribution'
    cont_k = (gt_k / gt_k['TOTAL']); cont_k.name = 'Contribution'

    return swc_final, cwc_s, cwc_k, cont_s, cont_k, gt_s, gt_k, sale, stock, grand, avail

# ══════════════════ EXCEL BUILDER ══════════════════
def _c(row, col, ws, val, bg, fg, sz=9, bold=False, h="center", ind=0, wrap=False):
    """Helper: set cell value + style in one shot."""
    cell = ws.cell(row=row, column=col, value=val)
    cell.font      = Font(bold=bold, size=sz, color=fg, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=h, vertical="center", wrap_text=wrap, indent=ind)
    cell.border    = Border(
        left=Side(style='thin', color="e5e7eb"),
        right=Side(style='thin', color="e5e7eb"),
        top=Side(style='thin', color="e5e7eb"),
        bottom=Side(style='thin', color="e5e7eb"),
    )
    return cell

def build_excel(swc_final, cwc_s, cwc_k, cont_s, cont_k, gt_s, gt_k, avail):
    wb = Workbook()

    # ── Pastel palette ──
    M_BG = ["dbeafe","dcfce7","fef9c3","ffedd5","fce7f3",
            "ede9fe","d1fae5","fee2e2","e0f2fe","d1fae5","fef9c3"]
    M_FG = ["1e40af","166534","854d0e","9a3412","9d174d",
            "5b21b6","065f46","991b1b","0c4a6e","065f46","854d0e"]
    C_BG = ["dbeafe","dcfce7","fef9c3","ffedd5","fce7f3",
            "ede9fe","d1fae5","fee2e2","e0f2fe","d1fae5","fef9c3"]
    C_FG = ["1e40af","166534","854d0e","9a3412","9d174d",
            "5b21b6","065f46","991b1b","0c4a6e","065f46","854d0e"]
    S_BG = ["eff6ff","f0fdf4","fefce8","fff7ed","fdf2f8",
            "f5f3ff","ecfdf5","fef2f2","f0f9ff","ecfdf5","fefce8"]
    S_FG = ["6b7280","6b7280","6b7280","6b7280","6b7280",
            "6b7280","6b7280","6b7280","6b7280","6b7280","6b7280"]

    HDR_BG = "f3f4f6"; HDR_FG = "111827"
    TITLE_BG = "1e3a5f"; TITLE_FG = "FFFFFF"
    CONT_BG = "f9fafb"; CONT_FG = "6b7280"
    GT_BG = "e8eaf6";   GT_FG = "1a237e"
    WHITE = "FFFFFF";    DGREY = "374151"

    # ════════════════════════════════════════
    # SWC SHEET
    # ════════════════════════════════════════
    ws = wb.active; ws.title = "SWC"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1565c0"

    all_cols = ["Row Labels"] + MONTHS + ["Total Sale", "Feb Closing Stk", "Sale Cont."]
    n = len(all_cols)

    ws.merge_cells(f"A1:{get_column_letter(n)}1")
    _c(1, 1, ws, "STORE WISE CONTRIBUTION (SWC)", TITLE_BG, TITLE_FG, sz=13, bold=True)
    ws.row_dimensions[1].height = 32

    mc_row = swc_final.loc['_mc']
    _c(2, 1, ws, "", CONT_BG, CONT_FG)
    for mi, col in enumerate(MONTHS):
        ci = mi + 2
        v = mc_row[col]
        _c(2, ci, ws, f"{v*100:.2f}%" if pd.notna(v) else "",
           M_BG[mi], M_FG[mi], sz=8, bold=True)
    _c(2, n-2, ws, "100.00%", "e8eaf6", "1a237e", sz=8, bold=True)
    _c(2, n-1, ws, "", CONT_BG, CONT_FG)
    _c(2, n,   ws, "", CONT_BG, CONT_FG)
    ws.row_dimensions[2].height = 17

    _c(3, 1, ws, "Store Name", HDR_BG, HDR_FG, sz=9, bold=True, h="left", ind=1)
    for mi, col in enumerate(MONTHS):
        ci = mi + 2
        label = MONTH_SHORT[mi]
        _c(3, ci, ws, label, M_BG[mi], M_FG[mi], sz=9, bold=True)
    _c(3, n-2, ws, "Total Sale",      "dbeafe", "1e40af", sz=9, bold=True)
    _c(3, n-1, ws, "Feb Closing Stk", "f3f4f6", "374151", sz=9, bold=True)
    _c(3, n,   ws, "Sale Cont.",       "dcfce7", "166534", sz=9, bold=True)
    ws.row_dimensions[3].height = 26

    data_rows = swc_final[~swc_final.index.isin(['_mc', 'Grand Total'])]
    for ri, (store, row) in enumerate(data_rows.iterrows(), 4):
        _c(ri, 1, ws, str(store), WHITE, DGREY, sz=9, h="left", ind=1)
        for mi, col in enumerate(MONTHS):
            ci = mi + 2
            v = row[col]
            val = int(v) if pd.notna(v) and v != 0 else "—"
            _c(ri, ci, ws, val, M_BG[mi], M_FG[mi], sz=9)
        ts = row['Total Sale']
        _c(ri, n-2, ws, int(ts) if pd.notna(ts) else "—", "dbeafe", "1e40af", sz=9, bold=True)
        v = row['Feb Closing Stk']
        _c(ri, n-1, ws, int(v) if pd.notna(v) and v != 0 else "—", WHITE, "6b7280", sz=9)
        v = row['Sale Cont.']
        _c(ri, n, ws, pct(v,4) if pd.notna(v) else "—", "dcfce7", "166534", sz=9, bold=True)
        ws.row_dimensions[ri].height = 18

    gr = len(data_rows) + 4
    gt_row = swc_final.loc['Grand Total']
    _c(gr, 1, ws, "Grand Total", GT_BG, GT_FG, sz=10, bold=True, h="left", ind=1)
    for mi, col in enumerate(MONTHS):
        ci = mi + 2
        v = gt_row[col]
        _c(gr, ci, ws, int(v) if pd.notna(v) and v != 0 else "—", GT_BG, GT_FG, sz=9, bold=True)
    _c(gr, n-2, ws, int(gt_row['Total Sale']),         GT_BG, GT_FG, sz=10, bold=True)
    v_stk = gt_row['Feb Closing Stk']
    _c(gr, n-1, ws, int(v_stk) if pd.notna(v_stk) else "—", GT_BG, GT_FG, sz=9, bold=True)
    _c(gr, n,   ws, "100.00%",                          GT_BG, GT_FG, sz=9,  bold=True)
    ws.row_dimensions[gr].height = 22

    ws.column_dimensions['A'].width = 22
    for ci in range(2, n+1):
        ws.column_dimensions[get_column_letter(ci)].width = 11
    ws.freeze_panes = "B4"

    # ════════════════════════════════════════
    # CWC SHEET
    # ════════════════════════════════════════
    ws2 = wb.create_sheet("CWC")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "0288d1"

    ncats = len(avail)
    total_cols = 1 + ncats * 2 + 2

    ws2.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    _c(1, 1, ws2, "CATEGORY WISE CONTRIBUTION (CWC)", TITLE_BG, TITLE_FG, sz=13, bold=True)
    ws2.row_dimensions[1].height = 32

    _c(2, 1, ws2, "Contribution", CONT_BG, CONT_FG, sz=8, bold=True, h="left", ind=1)
    ci = 2
    for i, cat in enumerate(avail):
        cbg = C_BG[i % len(C_BG)]; cfg = C_FG[i % len(C_FG)]
        sv = cont_s.get(cat, 0)
        kv = cont_k.get(cat, 0) if hasattr(cont_k, 'get') else 0
        _c(2, ci,   ws2, pct(sv,2) if pd.notna(sv) and sv!=0 else "—", cbg, cfg, sz=8, bold=True)
        _c(2, ci+1, ws2, pct(kv,2) if pd.notna(kv) and kv!=0 else "—", S_BG[i%len(S_BG)], S_FG[0], sz=8)
        ci += 2
    _c(2, ci,   ws2, "100%", "e8eaf6", "1a237e", sz=8, bold=True)
    _c(2, ci+1, ws2, "100%", "f9fafb", "6b7280", sz=8)
    ws2.row_dimensions[2].height = 17

    _c(3, 1, ws2, "Total", GT_BG, GT_FG, sz=9, bold=True, h="left", ind=1)
    ci = 2
    for i, cat in enumerate(avail):
        cbg = C_BG[i % len(C_BG)]; cfg = C_FG[i % len(C_FG)]
        sv = gt_s.get(cat, 0); kv = gt_k.get(cat, 0)
        _c(3, ci,   ws2, int(sv) if sv else "—", cbg, cfg, sz=9, bold=True)
        _c(3, ci+1, ws2, int(kv) if kv else "—", S_BG[i%len(S_BG)], S_FG[0], sz=9)
        ci += 2
    _c(3, ci,   ws2, int(gt_s['TOTAL']), GT_BG, GT_FG, sz=10, bold=True)
    _c(3, ci+1, ws2, int(gt_k['TOTAL']) if gt_k['TOTAL'] > 0 else "—", GT_BG, "6b7280", sz=9)
    ws2.row_dimensions[3].height = 20

    _c(4, 1, ws2, "", HDR_BG, HDR_FG)
    ci = 2
    _bdr = Border(left=Side(style='thin',color="e5e7eb"),right=Side(style='thin',color="e5e7eb"),
                  top=Side(style='thin',color="e5e7eb"),bottom=Side(style='thin',color="e5e7eb"))
    for i, cat in enumerate(avail):
        cbg = C_BG[i % len(C_BG)]; cfg = C_FG[i % len(C_FG)]
        ws2.merge_cells(start_row=4, start_column=ci, end_row=4, end_column=ci+1)
        _c(4, ci, ws2, cat, cbg, cfg, sz=9, bold=True, wrap=True)
        ws2.cell(row=4, column=ci+1).fill = PatternFill("solid", fgColor=cbg)
        ws2.cell(row=4, column=ci+1).border = _bdr
        ci += 2
    ws2.merge_cells(start_row=4, start_column=ci, end_row=4, end_column=ci+1)
    _c(4, ci, ws2, "TOTAL", "e8eaf6", "1a237e", sz=9, bold=True)
    ws2.cell(row=4, column=ci+1).fill = PatternFill("solid", fgColor="e8eaf6")
    ws2.cell(row=4, column=ci+1).border = _bdr
    ws2.row_dimensions[4].height = 24

    _c(5, 1, ws2, "Store Name", HDR_BG, HDR_FG, sz=9, bold=True)
    ci = 2
    for i in range(len(avail)):
        cbg = C_BG[i % len(C_BG)]; cfg = C_FG[i % len(C_FG)]
        _c(5, ci,   ws2, "Sale (MRP Value)",  cbg, cfg, sz=8, bold=True)
        _c(5, ci+1, ws2, "Stock (MRP Value)", S_BG[i%len(S_BG)], S_FG[0], sz=8, bold=True)
        ci += 2
    _c(5, ci,   ws2, "Sale (MRP Value)",  "e8eaf6", "1a237e", sz=8, bold=True)
    _c(5, ci+1, ws2, "Stock (MRP Value)", "f3f4f6", "6b7280", sz=8, bold=True)
    ws2.row_dimensions[5].height = 18

    stores_cwc = cwc_s.index.tolist()
    for ri, store in enumerate(stores_cwc, 6):
        _c(ri, 1, ws2, str(store), WHITE, DGREY, sz=9, h="left", ind=1)
        ci = 2
        for i, cat in enumerate(avail):
            cbg = C_BG[i % len(C_BG)]; cfg = C_FG[i % len(C_FG)]
            sv = cwc_s.loc[store, cat] if store in cwc_s.index else 0
            kv = cwc_k.loc[store, cat] if store in cwc_k.index else 0
            _c(ri, ci,   ws2, int(sv) if sv else "—", cbg, cfg, sz=9)
            _c(ri, ci+1, ws2, int(kv) if kv else "—", S_BG[i%len(S_BG)], S_FG[0], sz=9)
            ci += 2
        ts = cwc_s.loc[store, 'TOTAL'] if store in cwc_s.index else 0
        tk = cwc_k.loc[store, 'TOTAL'] if store in cwc_k.index else 0
        _c(ri, ci,   ws2, int(ts) if ts else "—", "dbeafe", "1e40af", sz=9, bold=True)
        _c(ri, ci+1, ws2, int(tk) if tk else "—", WHITE, "6b7280", sz=9)
        ws2.row_dimensions[ri].height = 18

    gr2 = len(stores_cwc) + 6
    _c(gr2, 1, ws2, "Grand Total", GT_BG, GT_FG, sz=10, bold=True, h="left", ind=1)
    ci = 2
    for i, cat in enumerate(avail):
        sv = gt_s.get(cat, 0); kv = gt_k.get(cat, 0)
        _c(gr2, ci,   ws2, int(sv) if sv else "—", GT_BG, GT_FG, sz=9, bold=True)
        _c(gr2, ci+1, ws2, int(kv) if kv else "—", GT_BG, "6b7280", sz=9, bold=True)
        ci += 2
    _c(gr2, ci,   ws2, int(gt_s['TOTAL']), GT_BG, GT_FG, sz=10, bold=True)
    _c(gr2, ci+1, ws2, int(gt_k['TOTAL']) if gt_k['TOTAL'] else "—", GT_BG, "6b7280", sz=9)
    ws2.row_dimensions[gr2].height = 22

    ws2.column_dimensions['A'].width = 22
    for ci in range(2, total_cols + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 14
    ws2.freeze_panes = "B6"

    out = BytesIO(); wb.save(out); out.seek(0)
    return out


# ══════════════════ SESSION STATE ══════════════════
for k, v in {"ready": False, "data": None}.items():
    if k not in st.session_state: st.session_state[k] = v

# ══════════════════ UI ══════════════════
st.markdown("""
<div class="hero">
  <div class="hero-badge">Sale Analyzer</div>
  <div class="hero-divider"></div>
  <div style="flex:1;min-width:0">
    <div style="display:flex;align-items:baseline;gap:.6rem;flex-wrap:wrap">
      <div class="hero-title">Sale Analyzer</div>
      <div class="hero-arrow">→</div>
      <div class="hero-sub-line">Store-wise Contribution &amp; Category-wise Contribution</div>
    </div>
    <div class="hero-sub">Upload Sale Report &nbsp;·&nbsp; Auto-generate Reports &nbsp;·&nbsp; Interactive Dashboard</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Upload ──
c1, c2, c3 = st.columns([1,2,1])
with c2:
    uploaded = st.file_uploader("📂 Drag & Drop Sale_Report.xlsx here", type=["xlsx","xls"], label_visibility="visible")
    if uploaded:
        b1,b2,b3 = st.columns([1,2,1])
        with b2:
            gen_clicked = st.button("⚡  Generate Reports + Dashboard", use_container_width=True)
        if gen_clicked:
            with st.spinner("Processing..."):
                result = process(uploaded)
                st.session_state.data  = result
                st.session_state.ready = True
            st.success("✅ Done! Reports Ready To Download.")

if not st.session_state.ready:
    st.markdown("""
    <div style="text-align:center;padding:5rem 0">
      <div style="font-size:3.5rem">📊</div>
      <div style="margin-top:1rem;font-size:1rem;color:#607d9b;font-weight:500">Upload Sale_Report.xlsx and click Generate</div>
      <div style="margin-top:.4rem;font-size:.82rem;color:#90a4c0">SWC + CWC Reports + Interactive Dashboard will be ready</div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ── Unpack ──
swc_final, cwc_s, cwc_k, cont_s, cont_k, gt_s, gt_k, sale, stock, grand, avail = st.session_state.data

data_swc = swc_final[~swc_final.index.isin(['_mc','Grand Total'])]
stores = data_swc.index.tolist()
total_closing = data_swc['Feb Closing Stk'].sum()
top_store = data_swc['Total Sale'].idxmax()
top_val = data_swc.loc[top_store,'Total Sale']

# ── KPIs ──
k1,k2,k3,k4 = st.columns(4)
for col, lbl, val, sub, icon in [
    (k1, "Total MRP Sale",   f"₹{fmt_inr(grand)}",        f"Apr'25 – Feb'26 · Include {len(stores)} Stores", "💰"),
    (k2, "Closing Stock",    f"₹{fmt_inr(total_closing)}", f"Feb 2026 · <span style='font-size:1rem;font-weight:800;color:#fff'>{int(stock[stock['Month']=='Feb Closing']['Quantity'].sum()) if 'Quantity' in stock.columns else ''} Pcs</span>", "📦"),
    (k3, "Top Store",        top_store.replace("SS, ",""), f"<span style='font-size:1rem;font-weight:800;color:#fff'>₹{fmt_inr(top_val)}</span> · Apr'25–Feb'26", "🏆"),
    (k4, "Categories",       str(len(avail)),              "Active Categories",             "🏷️"),
]:
    with col:
        st.markdown(f"""<div class="kpi-card">
          <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:.2rem">
            <div class="kpi-label">{lbl}</div>
            <span style="font-size:1.3rem;opacity:0.85">{icon}</span>
          </div>
          <div class="kpi-value">{val}</div>
          <div class="kpi-sub">{sub}</div>
        </div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ── Build HTML Dashboard ──
def build_html_dashboard(data_swc, cwc_s, gt_s, gt_k, avail, grand, stores, MONTHS, MONTH_SHORT, BLUE_SEQ, CAT_COLORS_LIGHT, fmt_inr, pct, chart_layout, stock=None):

    import json
    def clean(obj):
        if isinstance(obj, dict): return {k: clean(v) for k, v in obj.items()}
        elif isinstance(obj, (list, tuple)): return [clean(i) for i in obj]
        elif hasattr(obj, 'item'): return obj.item()
        elif hasattr(obj, 'tolist'): return obj.tolist()
        return obj

    _ctr = [0]
    def fig_div(fig):
        _ctr[0] += 1
        did = f"ch{_ctr[0]}"
        d = clean(fig.to_dict())
        j = json.dumps(d)
        return f'<div id="{did}" style="width:100%;overflow:hidden"></div><script>var _d{did}={j};Plotly.newPlot("{did}",_d{did}.data,_d{did}.layout,{{responsive:true,displayModeBar:false,displaylogo:false}});</script>'

    # KPIs
    total_closing = data_swc['Feb Closing Stk'].sum()
    total_closing_qty = int(stock[stock['Month']=='Feb Closing']['Quantity'].sum()) if stock is not None and 'Quantity' in stock.columns else 0
    top_store = data_swc['Total Sale'].idxmax().replace("SS, ","")
    top_val = data_swc['Total Sale'].max()

    # ── Chart 1: Monthly Bar ──
    monthly = data_swc[MONTHS].sum()
    f1 = go.Figure(go.Bar(
        x=MONTH_SHORT, y=monthly.values.tolist(),
        marker=dict(color=monthly.values.tolist(), colorscale=BLUE_SEQ, line=dict(width=0)),
        text=[f"₹{fmt_inr(v)}" for v in monthly.values],
        textposition='outside', textfont=dict(size=13, color='#1a0030'),
    ))
    f1.update_layout(**chart_layout(400, "Monthly MRP Sale — All Stores Combined (Apr'25 – Feb'26)"), bargap=0.3, yaxis_range=[0, float(monthly.max()) * 1.22])
    f1.update_layout(margin=dict(l=60, r=20, t=55, b=60))

    # ── Chart 2: Top 10 Stores ──
    top10 = data_swc['Total Sale'].nlargest(10).sort_values()
    f2 = go.Figure(go.Bar(
        x=top10.values.tolist(), y=top10.index.str.replace("SS, ","").tolist(),
        orientation='h',
        marker=dict(color=top10.values.tolist(), colorscale=BLUE_SEQ, line=dict(width=0)),
        text=[f"₹{fmt_inr(v)}" for v in top10.values],
        textposition='outside', textfont=dict(size=12, color='#1a0030'),
    ))
    f2.update_layout(**chart_layout(420, "Top 10 Stores by MRP Sale (Apr'25 – Feb'26)"), xaxis_range=[0, float(top10.max()) * 1.42])
    f2.update_layout(margin=dict(l=160, r=180, t=55, b=40))

    # ── Chart 3: Category Pie ──
    cat_vals = gt_s[avail]
    f3 = go.Figure(go.Pie(
        labels=list(avail), values=cat_vals.values.tolist(), hole=0.52,
        marker=dict(colors=CAT_COLORS_LIGHT[:len(avail)], line=dict(color='#fff', width=2)),
        textinfo='label+percent',
        textfont=dict(size=12, color='#1a0030'),
        insidetextfont=dict(size=11, color='#ffffff'),
    ))
    f3.update_layout(**chart_layout(400, "Category-wise Sale Contribution"),
        annotations=[dict(text=f"<b>₹{fmt_inr(grand)}</b>", x=0.5, y=0.5,
                          font=dict(size=13, color='#1a0030', family='Plus Jakarta Sans'), showarrow=False)])

    # ── Chart 4: SWC Store-wise Grouped Bar Chart ──
    bar_colors = ['#7b1fa2','#e91e63','#ff6f00','#1565c0','#2e7d32','#00838f',
                   '#f57f17','#6a1b9a','#c62828','#00695c']
    f5 = go.Figure()
    for i, store in enumerate(stores):
        row = data_swc.loc[store, MONTHS].fillna(0)
        f5.add_trace(go.Bar(
            x=MONTH_SHORT, y=row.values.tolist(), name=store.replace("SS, ",""),
            marker_color=bar_colors[i % len(bar_colors)],
            hovertemplate=f'<b>{store.replace("SS, "," ")}</b><br>%{{x}}: ₹%{{y:,.0f}}<extra></extra>'
        ))
    f5.update_layout(**chart_layout(500, "Store-wise Monthly Sale Comparison (SWC)", show_legend=True))
    f5.update_layout(barmode='group', margin=dict(l=60, r=180, t=55, b=60))

    # ── Chart 5: CWC Sale vs Stock ──
    cats = list(avail)
    sale_vals = [float(gt_s[c]) for c in cats]
    stk_vals  = [float(gt_k[c]) for c in cats]
    f6 = go.Figure()
    f6.add_trace(go.Bar(name='Sale (MRP)', y=cats, x=sale_vals, orientation='h',
        marker_color='#7b1fa2', marker_line_width=0,
        text=[f"₹{fmt_inr(v)}" for v in sale_vals],
        textposition='outside', textfont=dict(size=12, color='#4a0072')))
    f6.add_trace(go.Bar(name='Closing Stock', y=cats, x=stk_vals, orientation='h',
        marker_color='#ce93d8', marker_line_width=0,
        text=[f"₹{fmt_inr(v)}" for v in stk_vals],
        textposition='outside', textfont=dict(size=12, color='#6b21a8')))
    max_v = max(max(sale_vals), max(stk_vals))
    f6.update_layout(**chart_layout(560, "Total Sale vs Closing Stock by Category (CWC)"), barmode='group', bargap=0.25, bargroupgap=0.05, xaxis_range=[0, max_v * 1.45], legend_orientation='h', legend_y=1.08)
    f6.update_layout(margin=dict(l=200, r=200, t=65, b=40))

    # ── Chart 6: Stores per Category ──
    cat_store_count = (cwc_s[avail] > 0).sum().sort_values(ascending=True)
    f7 = go.Figure(go.Bar(
        x=cat_store_count.values.tolist(), y=cat_store_count.index.tolist(),
        orientation='h',
        marker=dict(color=cat_store_count.values.tolist(), colorscale=BLUE_SEQ, line=dict(width=0)),
        text=[f"{v} Stores" for v in cat_store_count.values],
        textposition='outside', textfont=dict(size=12, color='#1a0030'),
    ))
    f7.update_layout(**chart_layout(420, "Store Coverage by Category"), xaxis_range=[0, float(cat_store_count.max()) * 1.38])
    f7.update_layout(margin=dict(l=210, r=120, t=55, b=40))

    # ── Chart 7: Heatmap ──
    hm_data = cwc_s[avail].fillna(0)
    f4 = go.Figure(go.Heatmap(
        z=hm_data.values.tolist(), x=list(avail),
        y=hm_data.index.str.replace("SS, ","").tolist(),
        colorscale=[[0,'#fdf8ff'],[0.2,'#e9d8f8'],[0.5,'#c084fc'],[0.75,'#9333ea'],[1,'#581c87']],
        text=[[f"₹{fmt_inr(v)}" if v > 0 else "" for v in row] for row in hm_data.values.tolist()],
        texttemplate="%{text}", textfont=dict(size=9, color='#1a0030'),
        hoverongaps=False,
        colorbar=dict(title="Sale ₹", tickfont=dict(color='#1a0030')),
    ))
    f4.update_layout(paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
        font=dict(color="#1a0030", family="Inter", size=11),
        height=720, margin=dict(l=180, r=20, t=55, b=80),
        title=dict(text="<b>Sale Heatmap: Store × Category (MRP Value)</b>",
                   font=dict(color='#1a0030', size=15, family='Plus Jakarta Sans')),
        xaxis=dict(tickangle=-30, tickfont=dict(size=10, color='#2d0050')),
        yaxis=dict(tickfont=dict(size=10, color='#2d0050'), autorange='reversed'))

    # ── Heatmap Insight HTML ──
    _hm = cwc_s[avail].fillna(0)
    _hm_max_val  = _hm.max().max()
    _hm_max_cat  = _hm.max().idxmax()
    _hm_max_store = _hm[_hm_max_cat].idxmax().replace("SS, ","")
    _cat_totals  = _hm.sum().sort_values(ascending=False)
    _top_cats_str = "<br>".join([f"<b>{c}</b> — ₹{fmt_inr(int(v))}" for c,v in _cat_totals.head(3).items()])
    _store_totals = _hm.sum(axis=1).sort_values()
    _weakest = _store_totals.index[0].replace("SS, ","")
    _weakest_val = _store_totals.values[0]
    _best    = _store_totals.index[-1].replace("SS, ","")
    _best_val = _store_totals.values[-1]
    _zero_sc = {s.replace("SS, ",""): len([c for c in avail if _hm.loc[s,c]==0]) for s in _hm.index if len([c for c in avail if _hm.loc[s,c]==0]) >= 7}
    _zero_str = ", ".join([f"<b>{s}</b> ({n} categories zero)" for s,n in list(_zero_sc.items())[:3]]) if _zero_sc else f"<b>{_weakest}</b> — Weakest store · ₹{fmt_inr(int(_weakest_val))}"
    hm_insight_html = f"""<div style="background:#f8faff;border:1.5px solid #c7d7f9;border-radius:12px;padding:1.1rem 1.4rem;margin-top:.8rem">
      <div style="font-size:.65rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#1e40af;margin-bottom:.8rem;">🔥 HEATMAP — KEY INSIGHTS</div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:.8rem;">
        <div style="background:#f5f3ff;border-radius:8px;padding:.8rem 1rem;border-left:4px solid #7c3aed;">
          <div style="font-size:.72rem;font-weight:700;color:#4c1d95;margin-bottom:.3rem;">🏆 HIGHEST SALE COMBINATION</div>
          <div style="font-size:.9rem;font-weight:800;color:#1a0030;"><b>{_hm_max_store}</b> → {_hm_max_cat}</div>
          <div style="font-size:.82rem;color:#4c1d95;margin-top:.2rem;">₹{fmt_inr(int(_hm_max_val))} — Top performing store-category pair</div>
        </div>
        <div style="background:#fef2f2;border-radius:8px;padding:.8rem 1rem;border-left:4px solid #dc2626;">
          <div style="font-size:.72rem;font-weight:700;color:#991b1b;margin-bottom:.3rem;">⚠️ WEAK STORES (Many Zero Categories)</div>
          <div style="font-size:.82rem;color:#1a0030;">{_zero_str}</div>
        </div>
        <div style="background:#f0fdf4;border-radius:8px;padding:.8rem 1rem;border-left:4px solid #16a34a;">
          <div style="font-size:.72rem;font-weight:700;color:#166534;margin-bottom:.3rem;">📦 TOP CATEGORIES BY SALE</div>
          <div style="font-size:.82rem;color:#1a0030;">{_top_cats_str}</div>
        </div>
        <div style="background:#fefce8;border-radius:8px;padding:.8rem 1rem;border-left:4px solid #ca8a04;">
          <div style="font-size:.72rem;font-weight:700;color:#854d0e;margin-bottom:.3rem;">💡 HOW TO READ THIS HEATMAP</div>
          <div style="font-size:.82rem;color:#1a0030;">
            <b>Dark Purple</b> = High sale ✅<br>
            <b>Light Pink</b> = Low sale ⚠️<br>
            <b>Blank/—</b> = Zero sale ❌<br>
            Focus dark cells = your best opportunities
          </div>
        </div>
      </div>
      <div style="margin-top:.8rem;padding:.6rem .8rem;background:#fff;border-radius:8px;font-size:.78rem;color:#374151;">
        📌 <b>Best Store:</b> {_best} (₹{fmt_inr(int(_best_val))}) &nbsp;·&nbsp;
        <b>Weakest Store:</b> {_weakest} (₹{fmt_inr(int(_weakest_val))}) &nbsp;·&nbsp;
        <b>Top Category:</b> {_cat_totals.index[0]} (₹{fmt_inr(int(_cat_totals.values[0]))})
      </div>
    </div>"""

    # ── Chart 8: Deep Dive ──
    store_ranks = data_swc['Total Sale'].rank(ascending=False).astype(int)
    import json as _json
    store_data = {}
    for st in stores:
        row = data_swc.loc[st]
        monthly_v = [float(row[m]) if not pd.isna(row[m]) else 0 for m in MONTHS]
        cat_v = {}
        if st in cwc_s.index:
            for c in avail:
                v = cwc_s.loc[st, c] if c in cwc_s.columns else 0
                if not pd.isna(v) and v > 0:
                    cat_v[c] = float(v)
        _cqty = int(stock[(stock['Store Name']==st) & (stock['Month']=='Feb Closing')]['Quantity'].sum()) if stock is not None and 'Quantity' in stock.columns else 0
        store_data[st] = {
            'monthly': monthly_v,
            'total': float(row['Total Sale']),
            'closing': float(row['Feb Closing Stk']) if not pd.isna(row['Feb Closing Stk']) else 0,
            'closing_qty': _cqty,
            'cont': float(row['Sale Cont.']) if not pd.isna(row['Sale Cont.']) else 0,
            'rank': int(store_ranks[st]),
            'cats': cat_v
        }
    store_data_json = _json.dumps(store_data)

    # ── Performance Charts ──
    monthly_total = data_swc[MONTHS].sum()
    mom_total = monthly_total.pct_change() * 100
    apr_val = float(monthly_total.values[0])
    mom_vals_html = [float(v) for v in mom_total.values[1:]]
    colors_mom = ['#16a34a' if v >= 0 else '#dc2626' for v in mom_vals_html]
    mom_texts_all = [f"{v:+.1f}%" for v in mom_vals_html]
    f_mom = go.Figure(go.Bar(
        x=MONTH_SHORT[1:], y=mom_vals_html,
        marker=dict(color=colors_mom, line=dict(width=0)),
        text=mom_texts_all,
        textposition='outside', textfont=dict(size=12, color='#1a0030'),
    ))
    f_mom.add_annotation(x=MONTH_SHORT[1], y=max(mom_vals_html)*0.9,
        text=f'📌 Apr Start: ₹{fmt_inr(int(apr_val))}',
        showarrow=False, font=dict(size=11, color='#9c27b0', family='Inter'),
        bgcolor='rgba(243,232,255,0.9)', bordercolor='#9c27b0', borderwidth=1)
    f_mom.update_layout(paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
        font=dict(color="#1a0030", family="Inter", size=12), height=360, bargap=0.3,
        margin=dict(l=20, r=20, t=50, b=40),
        title=dict(text="<b>Month-on-Month Sale Growth (%) — All Stores Combined</b>",
                   font=dict(color='#1a0030', size=15, family='Plus Jakarta Sans')),
        xaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=11), linecolor='#ddd6fe'),
        yaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=11), linecolor='#ddd6fe',
                   zeroline=True, zerolinecolor='#9c27b0', zerolinewidth=2))

    top5 = data_swc['Total Sale'].nlargest(5).sort_values()
    f_top5 = go.Figure(go.Bar(
        x=top5.values.tolist(), y=top5.index.str.replace("SS, ","").tolist(), orientation='h',
        marker=dict(color='#16a34a', line=dict(width=0)),
        text=[f"₹{fmt_inr(v)}" for v in top5.values],
        textposition='outside', textfont=dict(size=12, color='#1a0030')))
    f_top5.update_layout(paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
        font=dict(color="#1a0030", family="Inter", size=12), height=300,
        margin=dict(l=10, r=160, t=40, b=20),
        title=dict(text="<b>Top 5 Stores</b>", font=dict(color='#1a0030', size=14, family='Plus Jakarta Sans')),
        xaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=10), linecolor='#ddd6fe', range=[0, float(top5.max())*1.4]),
        yaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=11), linecolor='#ddd6fe'))

    bot5 = data_swc['Total Sale'].nsmallest(5).sort_values(ascending=False)
    f_bot5 = go.Figure(go.Bar(
        x=bot5.values.tolist(), y=bot5.index.str.replace("SS, ","").tolist(), orientation='h',
        marker=dict(color='#dc2626', line=dict(width=0)),
        text=[f"₹{fmt_inr(v)}" for v in bot5.values],
        textposition='outside', textfont=dict(size=12, color='#1a0030')))
    f_bot5.update_layout(paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
        font=dict(color="#1a0030", family="Inter", size=12), height=300,
        margin=dict(l=10, r=160, t=40, b=20),
        title=dict(text="<b>Bottom 5 Stores</b>", font=dict(color='#1a0030', size=14, family='Plus Jakarta Sans')),
        xaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=10), linecolor='#ddd6fe', range=[0, float(bot5.max())*1.5]),
        yaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=11), linecolor='#ddd6fe'))

    # Zero sale + MoM table HTML
    zero_rows = ""
    for store in stores:
        row = data_swc.loc[store, MONTHS]
        zero_months = [MONTH_SHORT[i] for i, v in enumerate(row.values) if pd.isna(v) or v == 0]
        last = float(row.values[-1]) if not pd.isna(row.values[-1]) else 0.0
        prev = float(row.values[-2]) if not pd.isna(row.values[-2]) else 0.0
        growth = ((last - prev) / prev * 100) if prev > 0 else None
        if growth is None:
            arrow = "N/A"
        elif growth >= 0:
            arrow = f'<span style="color:#16a34a">▲ +{growth:.1f}%</span>'
        else:
            arrow = f'<span style="color:#dc2626">▼ {growth:.1f}%</span>'
        sale_months = [MONTH_SHORT[i] for i, v in enumerate(row.values) if not (pd.isna(v) or v == 0)]
        zero_rows += f"""<tr style="border-bottom:1px solid #f3f4f6">
            <td style="padding:.4rem .6rem;font-weight:600">{store.replace("SS, ","")}</td>
            <td style="padding:.4rem .6rem;color:#16a34a">{", ".join(sale_months) if sale_months else "—"}</td>
            <td style="padding:.4rem .6rem;color:#dc2626">{", ".join(zero_months) if zero_months else "—"}</td>
            <td style="padding:.4rem .6rem;text-align:center">{len(zero_months)}</td>
            <td style="padding:.4rem .6rem;text-align:center">{arrow}</td>
            <td style="padding:.4rem .6rem;text-align:right">₹{fmt_inr(data_swc.loc[store,"Total Sale"])}</td>
        </tr>"""
    zero_mom_html = f"""<table style="width:100%;border-collapse:collapse;font-size:.82rem">
        <thead><tr style="background:#f3f4f6">
            <th style="padding:.5rem .6rem;text-align:left">Store</th>
            <th style="padding:.5rem .6rem;text-align:left">Sale Months</th>
            <th style="padding:.5rem .6rem;text-align:left">Zero Sale Months</th>
            <th style="padding:.5rem .6rem;text-align:center">Count</th>
            <th style="padding:.5rem .6rem;text-align:center">Jan→Feb Growth</th>
            <th style="padding:.5rem .6rem;text-align:right">Total Sale</th>
        </tr></thead><tbody>{zero_rows}</tbody></table>"""

    # ── Inventory Intelligence Charts ──
    import pandas as _pd
    st_matrix = _pd.DataFrame(index=cwc_s.index, columns=list(avail))
    for _store in cwc_s.index:
        for _cat in avail:
            _s = float(cwc_s.loc[_store, _cat]) if _store in cwc_s.index else 0
            _k = float(cwc_k.loc[_store, _cat]) if _store in cwc_k.index else 0
            _t = _s + _k
            st_matrix.loc[_store, _cat] = round(_s/_t*100,1) if _t > 0 else 0
    st_matrix = st_matrix.astype(float)

    f_st = go.Figure(go.Heatmap(
        z=st_matrix.values.tolist(), x=list(avail),
        y=st_matrix.index.str.replace("SS, ","").tolist(),
        colorscale=[[0,'#fef2f2'],[0.3,'#fca5a5'],[0.6,'#fde68a'],[0.8,'#86efac'],[1,'#16a34a']],
        text=[[f"{v:.0f}%" if v > 0 else "—" for v in row] for row in st_matrix.values.tolist()],
        texttemplate="%{text}", textfont=dict(size=9, color='#1a0030'),
        hoverongaps=False, zmin=0, zmax=100,
        colorbar=dict(title="ST%", tickfont=dict(color='#1a0030', size=9), ticksuffix="%")))
    f_st.update_layout(paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(255,255,255,1)",
        font=dict(color="#1a0030", family="Inter", size=11), height=700,
        margin=dict(l=150, r=30, t=50, b=80),
        title=dict(text="<b>Sell-Through Rate (%) — Store × Category</b>",
                   font=dict(color='#1a0030', size=15, family='Plus Jakarta Sans')),
        xaxis=dict(tickangle=-30, tickfont=dict(size=10, color='#1a0030')),
        yaxis=dict(tickfont=dict(size=10, color='#1a0030'), autorange='reversed'))

    # Sell-through insights
    avg_st_val = float(st_matrix.replace(0, float('nan')).stack().mean())
    store_avg = st_matrix.replace(0, float('nan')).mean(axis=1).sort_values(ascending=False)
    best_s = store_avg[store_avg >= 60].index.str.replace("SS, ","").tolist()[:3]
    worst_s = store_avg[store_avg < 30].index.str.replace("SS, ","").tolist()[:3]

    restock_h = []
    for _store in cwc_s.index:
        for _cat in avail:
            _s = float(cwc_s.loc[_store, _cat]); _k = float(cwc_k.loc[_store, _cat]) if _store in cwc_k.index else 0
            if (_s+_k) > 0 and _s > 5000 and _s/(_s+_k)*100 >= 80:
                restock_h.append(f"{_store.replace('SS, ','')} → {_cat}")

    restock_cats = set([x.split(' → ')[1] for x in restock_h])
    cat_avg = st_matrix.replace(0, float('nan')).mean(axis=0).sort_values(ascending=False)
    truly_slow_c = [c for c in cat_avg[cat_avg < 30].index.tolist() if c not in restock_cats][:3]
    mixed_c = [c for c in cat_avg[cat_avg < 30].index.tolist() if c in restock_cats][:3]
    st_insight_html = f"""<div style="background:#f8faff;border:1.5px solid #c7d7f9;border-radius:12px;padding:1rem 1.2rem;margin-top:.8rem">
      <div style="font-size:.6rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#1e40af;margin-bottom:.7rem">📊 KEY INSIGHTS</div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:.7rem">
        <div style="background:#f0fdf4;border-radius:8px;padding:.7rem;border-left:4px solid #16a34a">
          <div style="font-size:.65rem;font-weight:700;color:#166534;margin-bottom:.3rem">🏆 BEST PERFORMING STORES</div>
          <div style="font-size:.78rem;color:#1a0030">{"<br>".join([f"<b>{s}</b> — Strong sell-through" for s in best_s]) if best_s else "N/A"}</div>
        </div>
        <div style="background:#fef2f2;border-radius:8px;padding:.7rem;border-left:4px solid #dc2626">
          <div style="font-size:.65rem;font-weight:700;color:#991b1b;margin-bottom:.3rem">⚠️ STORES NEEDING ATTENTION</div>
          <div style="font-size:.78rem;color:#1a0030">{"<br>".join([f"<b>{s}</b> — Low sell-through, overstocked" for s in worst_s]) if worst_s else "All stores OK"}</div>
        </div>
        <div style="background:#eff6ff;border-radius:8px;padding:.7rem;border-left:4px solid #1d4ed8">
          <div style="font-size:.65rem;font-weight:700;color:#1e40af;margin-bottom:.3rem">📦 URGENT RESTOCK NEEDED</div>
          <div style="font-size:.78rem;color:#1a0030">{"<br>".join(restock_h[:4]) if restock_h else "No urgent restock"}</div>
        </div>
        <div style="background:#fefce8;border-radius:8px;padding:.7rem;border-left:4px solid #ca8a04">
          <div style="font-size:.65rem;font-weight:700;color:#854d0e;margin-bottom:.3rem">🐌 SLOW MOVING CATEGORIES</div>
          <div style="font-size:.78rem;color:#1a0030">{"<br>".join([f"<b>{c}</b> — Reduce stock (slow across all stores)" for c in truly_slow_c]) if truly_slow_c else "No universally slow categories"}</div>
          {"<div style='font-size:.72rem;color:#ca8a04;margin-top:.3rem;font-style:italic'>" + ", ".join([f"<b>{c}</b>" for c in mixed_c]) + " — Mixed: slow in some stores, fast in others</div>" if mixed_c else ""}
        </div>
      </div>
      <div style="margin-top:.7rem;padding:.5rem .7rem;background:#fff;border-radius:6px;font-size:.75rem;color:#374151">
        📌 <b>Avg Sell-Through: {avg_st_val:.1f}%</b> &nbsp;·&nbsp;
        <span style="color:#16a34a;font-weight:600">Green ≥60%</span> = Good &nbsp;·&nbsp;
        <span style="color:#ca8a04;font-weight:600">Yellow 30–59%</span> = Average &nbsp;·&nbsp;
        <span style="color:#dc2626;font-weight:600">Red &lt;30%</span> = Slow mover
      </div>
    </div>"""

    # Fit score table
    fit_rows = ""
    for _store in cwc_s.index:
        for _cat in avail:
            _s = float(cwc_s.loc[_store,_cat]); _k = float(cwc_k.loc[_store,_cat]) if _store in cwc_k.index else 0
            _t = _s+_k
            if _t > 0:
                _r = _s/_t*100
                fit = "🟢 Strong" if _r>=60 else ("🟡 Average" if _r>=30 else "🔴 Weak")
                _sq = int(sale[sale['Store Name']==_store][sale['CATEGORY']==_cat]['Quantity'].sum()) if 'Quantity' in sale.columns else '—'
                _kq = int(stock[(stock['Store Name']==_store) & (stock['CATEGORY']==_cat) & (stock['Month']=='Feb Closing')]['Quantity'].sum()) if 'Quantity' in stock.columns else '—'
                fit_rows += f"<tr style='border-bottom:1px solid #f3f4f6'><td style='padding:.35rem .5rem;font-weight:600'>{_store.replace('SS, ','')}</td><td style='padding:.35rem .5rem'>{_cat}</td><td style='padding:.35rem .5rem;text-align:right'>₹{fmt_inr(_s)}</td><td style='padding:.35rem .5rem;text-align:center'>{_sq}</td><td style='padding:.35rem .5rem;text-align:right'>₹{fmt_inr(_k)}</td><td style='padding:.35rem .5rem;text-align:center'>{_kq}</td><td style='padding:.35rem .5rem;text-align:center'>{_r:.1f}%</td><td style='padding:.35rem .5rem;text-align:center'>{fit}</td></tr>"
    fit_html = f"""
    <div style="margin-bottom:.6rem;display:flex;gap:.5rem;flex-wrap:wrap;align-items:center">
      <span style="font-size:.72rem;font-weight:700;color:#6a1b9a;">Filter by Fit:</span>
      <button onclick="filterFit('All')" id="fitBtn_All" class="filt-btn filt-active" style="padding:.3rem .8rem;border-radius:6px;border:1.5px solid #c084fc;background:linear-gradient(135deg,#6a1b9a,#9c27b0);color:#fff;font-size:.75rem;font-weight:700;cursor:pointer">All</button>
      <button onclick="filterFit('Strong')" id="fitBtn_Strong" class="filt-btn" style="padding:.3rem .8rem;border-radius:6px;border:1.5px solid #86efac;background:#f0fdf4;color:#166534;font-size:.75rem;font-weight:700;cursor:pointer">🟢 Strong</button>
      <button onclick="filterFit('Average')" id="fitBtn_Average" class="filt-btn" style="padding:.3rem .8rem;border-radius:6px;border:1.5px solid #fde68a;background:#fefce8;color:#854d0e;font-size:.75rem;font-weight:700;cursor:pointer">🟡 Average</button>
      <button onclick="filterFit('Weak')" id="fitBtn_Weak" class="filt-btn" style="padding:.3rem .8rem;border-radius:6px;border:1.5px solid #fca5a5;background:#fef2f2;color:#991b1b;font-size:.75rem;font-weight:700;cursor:pointer">🔴 Weak</button>
    </div>
    <div style="max-height:400px;overflow-y:auto">
    <table id="fitTable" style="width:100%;border-collapse:collapse;font-size:.78rem">
        <thead style="position:sticky;top:0"><tr style="background:#f3f4f6">
            <th style="padding:.45rem .5rem;text-align:left">Store</th>
            <th style="padding:.45rem .5rem;text-align:left">Category</th>
            <th style="padding:.45rem .5rem;text-align:right">Sale (₹)</th>
            <th style="padding:.45rem .5rem;text-align:center">Sale Qty</th>
            <th style="padding:.45rem .5rem;text-align:right">Stock (₹)</th>
            <th style="padding:.45rem .5rem;text-align:center">Stock Qty</th>
            <th style="padding:.45rem .5rem;text-align:center">ST%</th>
            <th style="padding:.45rem .5rem;text-align:center">Fit</th>
        </tr></thead><tbody id="fitBody">{fit_rows}</tbody>
    </table></div>
    <script>
    function filterFit(val) {{
      document.querySelectorAll('.filt-btn').forEach(function(b) {{
        b.style.background = b.id === 'fitBtn_'+val ? 'linear-gradient(135deg,#6a1b9a,#9c27b0)' : (b.id.includes('Strong') ? '#f0fdf4' : b.id.includes('Average') ? '#fefce8' : b.id.includes('Weak') ? '#fef2f2' : '#f3f4f6');
        b.style.color = b.id === 'fitBtn_'+val ? '#fff' : (b.id.includes('Strong') ? '#166534' : b.id.includes('Average') ? '#854d0e' : b.id.includes('Weak') ? '#991b1b' : '#374151');
      }});
      var rows = document.querySelectorAll('#fitBody tr');
      rows.forEach(function(r) {{
        var fitCell = r.cells[7] ? r.cells[7].textContent : '';
        r.style.display = (val === 'All' || fitCell.includes(val)) ? '' : 'none';
      }});
    }}
    </script>"""

    # Dead stock table
    dead_rows = ""
    for _store in cwc_s.index:
        for _cat in avail:
            _s = float(cwc_s.loc[_store,_cat]); _k = float(cwc_k.loc[_store,_cat]) if _store in cwc_k.index else 0
            if _s == 0 and _k > 0:
                _dq = int(stock[(stock['Store Name']==_store) & (stock['CATEGORY']==_cat) & (stock['Month']=='Feb Closing')]['Quantity'].sum()) if 'Quantity' in stock.columns else '—'
                dead_rows += f"<tr style='border-bottom:1px solid #fecaca'><td style='padding:.35rem .5rem;font-weight:600'>{_store.replace('SS, ','')}</td><td style='padding:.35rem .5rem'>{_cat}</td><td style='padding:.35rem .5rem;text-align:right;color:#dc2626;font-weight:600'>₹{fmt_inr(_k)}</td><td style='padding:.35rem .5rem;text-align:center;color:#dc2626;font-weight:600'>{_dq} Pcs</td><td style='padding:.35rem .5rem;text-align:center'>⚠️ Transfer/Liquidate</td></tr>"
    _dead_val_total = sum([float(cwc_k.loc[_s,_c]) for _s in cwc_s.index for _c in avail if float(cwc_s.loc[_s,_c])==0 and float(cwc_k.loc[_s,_c] if _s in cwc_k.index else 0)>0])
    _dead_qty_total = sum([int(stock[(stock['Store Name']==_s) & (stock['CATEGORY']==_c) & (stock['Month']=='Feb Closing')]['Quantity'].sum()) for _s in cwc_s.index for _c in avail if float(cwc_s.loc[_s,_c])==0 and float(cwc_k.loc[_s,_c] if _s in cwc_k.index else 0)>0]) if 'Quantity' in stock.columns else 0
    _dead_count = dead_rows.count('<tr style=')
    dead_html = f"""<div style="background:#fef2f2;border:1px solid #fca5a5;border-radius:8px;padding:.5rem .8rem;font-size:.75rem;color:#991b1b;margin-bottom:.5rem">
        Dead Stock = Stock exists but ZERO sale. Immediate action needed!</div>
        <table style="width:100%;border-collapse:collapse;font-size:.78rem;max-height:350px;overflow-y:auto;display:block">
        <thead style="position:sticky;top:0"><tr style="background:#fef2f2">
            <th style="padding:.4rem .5rem;text-align:left">Store</th>
            <th style="padding:.4rem .5rem;text-align:left">Category</th>
            <th style="padding:.4rem .5rem;text-align:right">Dead Stock (₹)</th>
            <th style="padding:.4rem .5rem;text-align:center">Dead Stock Qty</th>
            <th style="padding:.4rem .5rem;text-align:center">Action</th>
        </tr></thead><tbody>{dead_rows if dead_rows else "<tr><td colspan=4 style='text-align:center;padding:1rem;color:#16a34a'>✅ No dead stock found!</td></tr>"}</tbody></table>
        <div style="background:#fef2f2;border:1px solid #fca5a5;border-radius:8px;padding:.6rem .8rem;margin-top:.5rem">
          <div style="font-size:.78rem;font-weight:700;color:#991b1b;margin-bottom:.3rem">⚠️ <b>{_dead_count} store-category combinations</b> store-category pairs have Dead Stock</div>
          <div style="display:flex;gap:1.5rem;font-size:.75rem;color:#7f1d1d">
            <span>💰 Total Dead Value: <b>₹{fmt_inr(int(_dead_val_total))}</b></span>
            <span>📦 Total Dead Qty: <b>{_dead_qty_total} Pcs</b></span>
          </div>
        </div>"""

    # Recommendations table
    rec_rows = ""
    for _store in cwc_s.index:
        for _cat in avail:
            _s = float(cwc_s.loc[_store,_cat]); _k = float(cwc_k.loc[_store,_cat]) if _store in cwc_k.index else 0
            _t = _s+_k
            if _t == 0: continue
            _r = _s/_t*100
            if _r >= 75 and _s > 5000: rec="📦 Increase Stock — High Demand"; pri="🔴 Urgent"
            elif _r >= 60 and _s > 0: rec="📦 Replenish Stock — Good Seller"; pri="🟡 Medium"
            elif _r < 20 and _k > 10000: rec="🔄 Transfer to Better Store"; pri="🔴 Urgent"
            elif _s == 0 and _k > 0: rec="❌ Remove Stock — No Sale"; pri="🔴 Urgent"
            elif _r < 30 and _k > 5000: rec="⬇️ Reduce Stock — Slow Mover"; pri="🟡 Medium"
            else: continue
            _rsq = int(sale[sale['Store Name']==_store][sale['CATEGORY']==_cat]['Quantity'].sum()) if 'Quantity' in sale.columns else '—'
            _rkq = int(stock[(stock['Store Name']==_store) & (stock['CATEGORY']==_cat) & (stock['Month']=='Feb Closing')]['Quantity'].sum()) if 'Quantity' in stock.columns else '—'
            rec_rows += f"<tr style='border-bottom:1px solid #f3f4f6'><td style='padding:.35rem .5rem;font-weight:600'>{_store.replace('SS, ','')}</td><td style='padding:.35rem .5rem'>{_cat}</td><td style='padding:.35rem .5rem;text-align:right'>₹{fmt_inr(_s)}</td><td style='padding:.35rem .5rem;text-align:center'>{_rsq}</td><td style='padding:.35rem .5rem;text-align:right'>₹{fmt_inr(_k)}</td><td style='padding:.35rem .5rem;text-align:center'>{_rkq}</td><td style='padding:.35rem .5rem;text-align:center'>{_r:.1f}%</td><td style='padding:.35rem .5rem'>{rec}</td><td style='padding:.35rem .5rem;text-align:center'>{pri}</td></tr>"
    rec_html = f"""
    <div style="margin-bottom:.6rem;display:flex;gap:.5rem;flex-wrap:wrap;align-items:center">
      <span style="font-size:.72rem;font-weight:700;color:#6a1b9a;">Filter by Priority:</span>
      <button onclick="filterRec('All')" id="recBtn_All" class="rec-btn rec-active" style="padding:.3rem .8rem;border-radius:6px;border:1.5px solid #c084fc;background:linear-gradient(135deg,#6a1b9a,#9c27b0);color:#fff;font-size:.75rem;font-weight:700;cursor:pointer">All</button>
      <button onclick="filterRec('Urgent')" id="recBtn_Urgent" class="rec-btn" style="padding:.3rem .8rem;border-radius:6px;border:1.5px solid #fca5a5;background:#fef2f2;color:#991b1b;font-size:.75rem;font-weight:700;cursor:pointer">🔴 Urgent</button>
      <button onclick="filterRec('Medium')" id="recBtn_Medium" class="rec-btn" style="padding:.3rem .8rem;border-radius:6px;border:1.5px solid #fde68a;background:#fefce8;color:#854d0e;font-size:.75rem;font-weight:700;cursor:pointer">🟡 Medium</button>
    </div>
    <div style="max-height:420px;overflow-y:auto">
    <table id="recTable" style="width:100%;border-collapse:collapse;font-size:.78rem">
        <thead style="position:sticky;top:0"><tr style="background:#f3f4f6">
            <th style="padding:.45rem .5rem;text-align:left">Store</th>
            <th style="padding:.45rem .5rem;text-align:left">Category</th>
            <th style="padding:.45rem .5rem;text-align:right">Sale (₹)</th>
            <th style="padding:.45rem .5rem;text-align:center">Sale Qty</th>
            <th style="padding:.45rem .5rem;text-align:right">Stock (₹)</th>
            <th style="padding:.45rem .5rem;text-align:center">Stock Qty</th>
            <th style="padding:.45rem .5rem;text-align:center">ST%</th>
            <th style="padding:.45rem .5rem;text-align:left">Recommendation</th>
            <th style="padding:.45rem .5rem;text-align:center">Priority</th>
        </tr></thead><tbody id="recBody">{rec_rows}</tbody>
    </table></div>
    <script>
    function filterRec(val) {{
      document.querySelectorAll('.rec-btn').forEach(function(b) {{
        b.style.background = b.id === 'recBtn_'+val ? 'linear-gradient(135deg,#6a1b9a,#9c27b0)' : (b.id.includes('Urgent') ? '#fef2f2' : b.id.includes('Medium') ? '#fefce8' : '#f3f4f6');
        b.style.color = b.id === 'recBtn_'+val ? '#fff' : (b.id.includes('Urgent') ? '#991b1b' : b.id.includes('Medium') ? '#854d0e' : '#374151');
      }});
      var rows = document.querySelectorAll('#recBody tr');
      rows.forEach(function(r) {{
        var priCell = r.cells[8] ? r.cells[8].textContent : '';
        r.style.display = (val === 'All' || priCell.includes(val)) ? '' : 'none';
      }});
    }}
    </script>"""
    month_short_json = _json.dumps(MONTH_SHORT)
    cat_colors_json = _json.dumps(CAT_COLORS_LIGHT)
    avail_json = _json.dumps(list(avail))
    stores_clean = [s.replace("SS, ","") for s in stores]
    stores_json = _json.dumps(stores)
    stores_clean_json = _json.dumps(stores_clean)
    total_stores = len(stores)

    deep_dive_html = f"""
    <div style="display:flex;align-items:center;gap:1rem;margin-bottom:1rem">
      <div style="flex:1">
        <label style="font-size:.75rem;font-weight:700;color:#6a1b9a;letter-spacing:1px;text-transform:uppercase;display:block;margin-bottom:.4rem">Select Store</label>
        <select id="ddStore" onchange="updateDeepDive()" style="width:100%;padding:.6rem 1rem;border:1.5px solid #c084fc;border-radius:10px;font-size:.9rem;color:#1a0030;background:#fff;outline:none">
          {''.join(f'<option value="{s}">{s.replace("SS, ","")}</option>' for s in stores)}
        </select>
      </div>
      <div id="ddRankCard" style="background:linear-gradient(135deg,#6a1b9a,#9c27b0);border-radius:12px;padding:.8rem 1.3rem;display:flex;align-items:center;gap:.8rem;min-width:220px;box-shadow:0 4px 14px rgba(106,27,154,.3)">
        <div style="background:rgba(255,255,255,.2);border-radius:50%;width:48px;height:48px;min-width:48px;display:flex;align-items:center;justify-content:center;border:2px solid rgba(255,255,255,.35)">
          <span id="ddRankNum" style="color:#fff;font-size:1.15rem;font-weight:800">#1</span>
        </div>
        <div>
          <div style="color:rgba(255,255,255,.65);font-size:.55rem;letter-spacing:2px;text-transform:uppercase;font-weight:700">Rank Among All</div>
          <div id="ddStoreName" style="color:#fff;font-size:.9rem;font-weight:700"></div>
          <div style="color:rgba(255,255,255,.65);font-size:.7rem">Out of {total_stores} stores</div>
        </div>
      </div>
    </div>
    <div class="kpi-row" style="grid-template-columns:repeat(3,1fr);margin-bottom:1rem">
      <div class="kpi"><div class="kpi-label">Total MRP Sale</div><div class="kpi-value" id="ddSale">—</div><div class="kpi-sub">Apr'25 → Feb'26 Total</div></div>
      <div class="kpi"><div class="kpi-label">Feb Closing Stock</div><div class="kpi-value" id="ddStock">—</div><div class="kpi-sub">Feb 2026 Closing</div></div>
      <div class="kpi"><div class="kpi-label">Sale Contribution</div><div class="kpi-value" id="ddCont">—</div><div class="kpi-sub">Of Total Sale</div></div>
    </div>
    <div class="two">
      <div class="card"><div class="sec">📅 Monthly Sale</div><div id="ddMonthChart" style="width:100%"></div></div>
      <div class="card"><div class="sec">🏷️ Category Mix</div><div id="ddCatChart" style="width:100%"></div></div>
    </div>
    <script>
    var _storeData = {store_data_json};
    var _months = {month_short_json};
    var _catColors = {cat_colors_json};
    var _avail = {avail_json};
    var _storesRaw = {stores_json};
    var _BLUE_SEQ = {_json.dumps(BLUE_SEQ)};

    function fmtInr(v) {{
      if (!v || v==0) return '—';
      v = Math.round(v);
      var s = Math.abs(v).toString(), pre = v<0?'-':'', last3 = s.slice(-3), rest = s.slice(0,-3), groups = [];
      while(rest.length>2){{groups.unshift(rest.slice(-2));rest=rest.slice(0,-2);}}
      if(rest) groups.unshift(rest);
      return '₹' + pre + groups.join(',') + ',' + last3;
    }}

    function updateDeepDive() {{
      var sel = document.getElementById('ddStore').value;
      var d = _storeData[sel];
      if(!d) return;
      var nm = sel.replace('SS, ','');
      document.getElementById('ddRankNum').textContent = '#'+d.rank;
      document.getElementById('ddStoreName').textContent = nm;
      document.getElementById('ddSale').textContent = fmtInr(d.total);
      document.getElementById('ddStock').innerHTML = fmtInr(d.closing) + (d.closing_qty ? "<div style='font-size:.75rem;font-weight:700;color:rgba(255,255,255,.85);margin-top:.2rem'>" + d.closing_qty.toLocaleString('en-IN') + " Pcs</div>" : '');
      document.getElementById('ddCont').textContent = (d.cont*100).toFixed(4)+'%';

      var mvals = d.monthly;
      var mmax = Math.max(...mvals) * 1.22;
      Plotly.newPlot('ddMonthChart', [{{
        type:'bar', x:_months, y:mvals,
        marker:{{color:mvals, colorscale:_BLUE_SEQ, line:{{width:0}}}},
        text:mvals.map(v=>v>0?fmtInr(v):''), textposition:'outside',
        textfont:{{size:11,color:'#1a0030'}}
      }}], {{
        paper_bgcolor:'rgba(255,255,255,1)', plot_bgcolor:'rgba(245,240,255,0.5)',
        height:280, margin:{{l:55,r:10,t:30,b:40}},
        font:{{color:'#1a0030',size:11}}, bargap:0.3,
        xaxis:{{gridcolor:'#ede9fe',tickfont:{{color:'#1a0030',size:10}}}},
        yaxis:{{gridcolor:'#ede9fe',tickfont:{{color:'#1a0030',size:10}},range:[0,mmax]}}
      }}, {{responsive:true,displayModeBar:false,displaylogo:false}});

      var cats = Object.keys(d.cats), vals = Object.values(d.cats);
      if(cats.length > 0) {{
        Plotly.newPlot('ddCatChart', [{{
          type:'pie', labels:cats, values:vals, hole:0.5,
          marker:{{colors:_catColors.slice(0,cats.length), line:{{color:'#fff',width:2}}}},
          textinfo:'label+percent',
          textfont:{{size:11,color:'#1a0030'}},
          insidetextfont:{{size:10,color:'#ffffff'}}
        }}], {{
          paper_bgcolor:'rgba(255,255,255,1)',
          height:280, margin:{{l:55,r:10,t:30,b:10}},
          font:{{color:'#1a0030',size:11}},
          showlegend:false
        }}, {{responsive:true,displayModeBar:false,displaylogo:false}});
      }}
    }}
    setTimeout(updateDeepDive, 300);
    </script>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Sale Analyzer — Dashboard</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js" charset="utf-8"></script>
<style>
  *{{margin:0;padding:0;box-sizing:border-box;font-family:'Segoe UI',Arial,sans-serif}}
  body{{background:#f4f0ff;color:#1a0030;overflow-x:hidden}}
  .navbar{{background:linear-gradient(90deg,#3a0068,#6a1b9a,#9c27b0);
    padding:.75rem 2rem;display:flex;align-items:center;gap:1rem;
    box-shadow:0 4px 16px rgba(106,27,154,.3);position:sticky;top:0;z-index:99}}
  .badge{{background:rgba(255,255,255,.2);border:1.5px solid rgba(255,255,255,.4);
    color:#fff;font-size:.55rem;font-weight:700;letter-spacing:2px;
    text-transform:uppercase;padding:4px 12px;border-radius:20px;white-space:nowrap}}
  .nav-title{{font-size:1.05rem;font-weight:800;color:#fff}}
  .nav-hint{{font-size:.62rem;color:rgba(255,255,255,.55);margin-top:.1rem}}
  .wrap{{width:100%;padding:1.2rem 1.5rem}}
  .kpi-row{{display:grid;grid-template-columns:repeat(4,1fr);gap:1rem;margin-bottom:1.4rem}}
  .kpi{{background:linear-gradient(135deg,#6a1b9a,#9c27b0);border-radius:14px;
    padding:1.1rem 1.3rem;box-shadow:0 4px 16px rgba(106,27,154,.3)}}
  .kpi-label{{font-size:.58rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;
    color:rgba(255,255,255,.7);margin-bottom:.4rem}}
  .kpi-value{{font-size:1.5rem;font-weight:800;color:#fff;line-height:1.1}}
  .kpi-sub{{font-size:.72rem;color:rgba(255,255,255,.65);margin-top:.25rem}}
  .tab-bar{{display:flex;gap:.5rem;margin-bottom:1.2rem;flex-wrap:wrap}}
  .tab{{padding:.55rem 1.1rem;border-radius:8px;border:1.5px solid #ddd6fe;
    background:#fff;color:#4c1d95;font-size:.82rem;font-weight:600;
    cursor:pointer;transition:all .18s;white-space:nowrap}}
  .tab:hover{{background:#f3e5f5}}
  .tab.active{{background:linear-gradient(135deg,#6a1b9a,#9c27b0);color:#fff;border-color:transparent}}
  .section{{display:none}}.section.active{{display:block}}
  .card{{background:#fff;border-radius:14px;padding:1.2rem 1.2rem;
    box-shadow:0 2px 12px rgba(106,27,154,.08);margin-bottom:1.2rem;overflow:visible}}
  .sec{{font-size:.6rem;font-weight:700;letter-spacing:2.5px;text-transform:uppercase;
    color:#6a1b9a;border-bottom:2px solid #ddd6fe;padding-bottom:.4rem;margin-bottom:.9rem}}
  .two{{display:grid;grid-template-columns:1fr 1fr;gap:1.2rem;margin-bottom:1.2rem}}
  .footer{{text-align:center;padding:1.5rem;color:#9575cd;font-size:.75rem}}
  @media(max-width:900px){{.kpi-row{{grid-template-columns:1fr 1fr}}.two{{grid-template-columns:1fr}}}}
</style>
</head>
<body>
<div class="navbar">
  <div class="badge">Sale Analyzer</div>
  <div style="width:1px;height:26px;background:rgba(255,255,255,.22)"></div>
  <div>
    <div class="nav-title">Sale Analyzer &nbsp;→&nbsp; <span style="color:#e8c8ff">Store-wise &amp; Category-wise Contribution</span></div>
    <div class="nav-hint">Interactive Dashboard &nbsp;·&nbsp; {pd.Timestamp.now().strftime('%d %b %Y')}</div>
  </div>
</div>
<div class="wrap">
  <div class="kpi-row">
    <div class="kpi"><div class="kpi-label">Total MRP Sale</div><div class="kpi-value">₹{fmt_inr(grand)}</div><div class="kpi-sub">Apr'25 – Feb'26 · Include {len(stores)} Stores</div></div>
    <div class="kpi"><div class="kpi-label">Closing Stock</div><div class="kpi-value">₹{fmt_inr(total_closing)}</div><div class="kpi-sub">Feb 2026 Closing</div><div style="font-size:.8rem;font-weight:700;color:rgba(255,255,255,.85);margin-top:.2rem">{total_closing_qty:,} Pcs</div></div>
    <div class="kpi"><div class="kpi-label">Top Store 🏆</div><div class="kpi-value" style="font-size:1.05rem">{top_store}</div><div class="kpi-sub">₹{fmt_inr(top_val)} · #1 of {len(stores)}</div></div>
    <div class="kpi"><div class="kpi-label">Categories</div><div class="kpi-value">{len(avail)}</div><div class="kpi-sub">Active Categories</div></div>
  </div>

  <div class="tab-bar">
    <div class="tab active" onclick="showTab('ov',this)">📈 Overview</div>
    <div class="tab" onclick="showTab('swc',this)">🏪 Store-wise (SWC)</div>
    <div class="tab" onclick="showTab('cwc',this)">📦 Category-wise (CWC)</div>
    <div class="tab" onclick="showTab('hm',this)">🔥 Heatmap</div>
    <div class="tab" onclick="showTab('dd',this)">🔍 Store Deep Dive</div>
    <div class="tab" onclick="showTab('perf',this)">📊 Performance</div>
    <div class="tab" onclick="showTab('inv',this)">📦 Inventory Intelligence</div>
  </div>

  <div id="ov" class="section active">
    <div class="card"><div class="sec">📈 Monthly Sale Trend</div>{fig_div(f1)}</div>
    <div class="two">
      <div class="card"><div class="sec">🏆 Top 10 Stores</div>{fig_div(f2)}</div>
      <div class="card"><div class="sec">🏷️ Category Mix</div>{fig_div(f3)}</div>
    </div>
  </div>

  <div id="swc" class="section">
    <div class="card"><div class="sec">🏪 Store-wise Monthly Sale Trend</div>{fig_div(f5)}</div>
  </div>

  <div id="cwc" class="section">
    <div class="card"><div class="sec">📦 Total Sale vs Closing Stock by Category</div>{fig_div(f6)}</div>
    <div class="card"><div class="sec">🏪 Store Coverage by Category</div>{fig_div(f7)}</div>
  </div>

  <div id="hm" class="section">
    <div class="card"><div class="sec">🔥 Sale Heatmap: Store × Category</div>{fig_div(f4)}{hm_insight_html}</div>
  </div>

  <div id="dd" class="section">
    <div class="card"><div class="sec">🔍 Store Deep Dive</div>{deep_dive_html}</div>
  </div>

  <div id="perf" class="section">
    <div class="card"><div class="sec">📊 Month-on-Month Growth</div>{fig_div(f_mom)}</div>
    <div class="two">
      <div class="card"><div class="sec">🏆 Top 5 Stores</div>{fig_div(f_top5)}</div>
      <div class="card"><div class="sec">🔴 Bottom 5 Stores</div>{fig_div(f_bot5)}</div>
    </div>
    <div class="card"><div class="sec">❌ Zero Sale Months & MoM Growth</div>{zero_mom_html}</div>
  </div>

  <div id="inv" class="section">
    <div class="card"><div class="sec">📊 Sell-Through Rate — Store × Category</div>{fig_div(f_st)}{st_insight_html}</div>
    <div class="two">
      <div class="card"><div class="sec">🟢 Category-Store Fit Score</div>{fit_html}</div>
      <div class="card"><div class="sec">⚠️ Dead Stock Alert</div>{dead_html}</div>
    </div>
    <div class="card"><div class="sec">📋 Stock Recommendations</div>{rec_html}</div>
  </div>

  <div class="footer">Sale Analyzer &nbsp;·&nbsp; Generated {pd.Timestamp.now().strftime('%d %b %Y %I:%M %p')}</div>
</div>
<script>
function showTab(id, el) {{
  document.querySelectorAll('.section').forEach(s => s.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  el.classList.add('active');
  setTimeout(function() {{
    window.dispatchEvent(new Event('resize'));
    var section = document.getElementById(id);
    var plots = section.querySelectorAll('[id^="ch"]');
    plots.forEach(function(p) {{
      if(p.id && window.Plotly) {{
        try {{ Plotly.Plots.resize(p.id); }} catch(e) {{}}
      }}
    }});
  }}, 200);
  if(id==='dd') setTimeout(updateDeepDive, 200);
}}
window.addEventListener('load', function() {{
  var tabs = ['ov','swc','cwc','hm','dd','perf','inv'];
  var activeTab = 'ov';
  tabs.forEach(function(id) {{
    if(id !== activeTab) {{
      var s = document.getElementById(id);
      if(s) {{
        s.style.display = 'block';
        s.style.visibility = 'hidden';
        s.style.position = 'absolute';
      }}
    }}
  }});
  setTimeout(function() {{
    window.dispatchEvent(new Event('resize'));
    tabs.forEach(function(id) {{
      if(id !== activeTab) {{
        var s = document.getElementById(id);
        if(s) {{
          s.style.display = '';
          s.style.visibility = '';
          s.style.position = '';
          s.classList.remove('active');
        }}
        var plots = document.querySelectorAll('#' + id + ' [id^="ch"]');
        plots.forEach(function(p) {{
          if(p.id && window.Plotly) {{
            try {{ Plotly.Plots.resize(p.id); }} catch(e) {{}}
          }}
        }});
      }}
    }});
  }}, 800);
}});
</script>
</body></html>"""
    return html.encode('utf-8')

# ── Download Buttons ──
excel = build_excel(swc_final, cwc_s, cwc_k, cont_s, cont_k, gt_s, gt_k, avail)
html_dash = build_html_dashboard(data_swc, cwc_s, gt_s, gt_k, avail, grand, stores, MONTHS, MONTH_SHORT, BLUE_SEQ, CAT_COLORS_LIGHT, fmt_inr, pct, chart_layout, stock=stock)

# ── Build ZIP ──
import zipfile, io as _io
excel_bytes = excel.getvalue() if hasattr(excel, 'getvalue') else (excel.read() if hasattr(excel, 'read') else excel)
zip_buf = _io.BytesIO()
with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
    zf.writestr("SS_Review_Report.xlsx", excel_bytes)
    zf.writestr("SS_Dashboard.html", html_dash)
zip_buf.seek(0)
zip_data = zip_buf.read()

dl1, dl2, dl3 = st.columns(3)
with dl1:
    st.download_button("📥  Excel Report (SWC + CWC)",
        data=excel, file_name="SS_Review_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True)
with dl2:
    st.download_button("🌐  Interactive Dashboard (HTML)",
        data=html_dash, file_name="SS_Dashboard.html",
        mime="text/html", use_container_width=True)
with dl3:
    st.download_button("📦  Download Both (ZIP)",
        data=zip_data, file_name="SS_Complete_Report.zip",
        mime="application/zip", use_container_width=True)

st.markdown("""<div style="background:#f3e5f5;border:1px solid #ce93d8;border-radius:10px;
    padding:.55rem 1rem;font-size:.75rem;color:#4a0072;margin:.4rem 0 0 0">
  💡 <b>ZIP contains both files:</b> Excel Report + Interactive HTML Dashboard &nbsp;·&nbsp;
  Open HTML file in browser for interactive charts &nbsp;·&nbsp;
  Ctrl+P → Save as PDF for print version
</div>""", unsafe_allow_html=True)

# ══════════════════ TABS ══════════════════
t1, t2, t3, t4, t5, t6, t7, t8 = st.tabs(["📈 Overview", "🏪 Store-wise (SWC)", "📦 Category-wise (CWC)", "🔥 Heatmap", "🔍 Store Deep Dive", "📊 Performance", "📦 Inventory Intelligence", "🤖 AI Strategy Summary"])

with t1:
    st.markdown('<div class="section-title">📈 Monthly Sale Trend</div>', unsafe_allow_html=True)
    monthly = data_swc[MONTHS].sum()
    best_idx = monthly.values.argmax()
    worst_idx = monthly.values.argmin()
    bar_colors = ['#9c27b0' if i not in [best_idx, worst_idx] else ('#16a34a' if i == best_idx else '#dc2626') for i in range(len(monthly))]
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=MONTH_SHORT, y=monthly.values,
        marker=dict(color=bar_colors, line=dict(width=0)),
        text=[f"₹{fmt_inr(v)}" for v in monthly.values],
        textposition='outside',
        textfont=dict(size=14, color='#1a0030', family='Inter'),
        hovertemplate='<b>%{x}</b><br>Sale: ₹%{y:,.0f}<extra></extra>',
    ))
    fig.update_layout(**chart_layout(380, "Monthly MRP Sale — All Stores Combined (Apr'25 – Feb'26)"),
        bargap=0.3,
        yaxis_range=[0, monthly.max() * 1.22],
        annotations=[
            dict(x=MONTH_SHORT[best_idx], y=monthly.values[best_idx]*1.18,
                 text="🏆 Best", showarrow=False, font=dict(color='#16a34a', size=11, family='Inter')),
            dict(x=MONTH_SHORT[worst_idx], y=monthly.values[worst_idx]*1.18,
                 text="⬇ Lowest", showarrow=False, font=dict(color='#dc2626', size=11, family='Inter')),
        ])
    st.plotly_chart(fig, use_container_width=True)

    avg_monthly = monthly.mean()
    feb_growth = ((monthly.values[-1] - monthly.values[0]) / monthly.values[0] * 100) if monthly.values[0] > 0 else 0
    ins1, ins2, ins3, ins4 = st.columns(4)
    for col, lbl, val, sub, icon in [
        (ins1, "Best Month",   f"{MONTH_SHORT[best_idx]}", f"₹{fmt_inr(monthly.values[best_idx])}", "📈"),
        (ins2, "Lowest Month", f"{MONTH_SHORT[worst_idx]}", f"₹{fmt_inr(monthly.values[worst_idx])}", "📉"),
        (ins3, "Avg Monthly",  f"₹{fmt_inr(int(avg_monthly))}", "Per month average", "📊"),
        (ins4, "Growth",       f"{feb_growth:+.1f}%", "Apr vs Feb trend", "🚀"),
    ]:
        with col:
            st.markdown(f"""<div style="background:#fff;border-radius:12px;padding:.9rem 1.1rem;
                box-shadow:0 2px 10px rgba(106,27,154,.1);border-left:4px solid #9c27b0;">
                <div style="font-size:.58rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#6a1b9a;">{icon} {lbl}</div>
                <div style="font-size:1.3rem;font-weight:800;color:#1a0030;margin:.25rem 0 .1rem">{val}</div>
                <div style="font-size:.72rem;color:#607d9b;">{sub}</div>
            </div>""", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    ca, cb = st.columns(2)
    with ca:
        st.markdown('<div class="section-title">🏆 Top 10 Stores</div>', unsafe_allow_html=True)
        top10 = data_swc['Total Sale'].nlargest(10).sort_values()
        fig2 = go.Figure(go.Bar(
            x=top10.values, y=top10.index.str.replace("SS, ",""),
            orientation='h',
            marker=dict(color=top10.values, colorscale=BLUE_SEQ, line=dict(width=0)),
            text=[f"₹{fmt_inr(v)}" for v in top10.values],
            textposition='outside',
            textfont=dict(size=13, color='#1a0030', family='Inter'),
            hovertemplate='<b>%{y}</b><br>₹%{x:,.0f}<extra></extra>',
        ))
        fig2.update_layout(**chart_layout(400, "Top 10 Stores by MRP Sale (Apr'25 – Feb'26)"),
            xaxis_range=[0, top10.max() * 1.35])
        st.plotly_chart(fig2, use_container_width=True)

    with cb:
        st.markdown('<div class="section-title">🏷️ Category Mix</div>', unsafe_allow_html=True)
        cat_vals = gt_s[avail]
        fig3 = go.Figure(go.Pie(
            labels=avail, values=cat_vals.values, hole=0.52,
            marker=dict(colors=CAT_COLORS_LIGHT[:len(avail)], line=dict(color='#ffffff', width=3)),
            textinfo='label+percent',
            textfont=dict(size=13, color='#1a0030', family='Inter'),
            insidetextfont=dict(size=12, color='#ffffff'),
            textposition='auto',
            hovertemplate='<b>%{label}</b><br>₹%{value:,.0f}<br>%{percent}<extra></extra>',
        ))
        fig3.update_layout(**chart_layout(400, "Category-wise Sale Contribution"),
            annotations=[dict(text=f"<b>₹{fmt_inr(grand)}</b>", x=0.5, y=0.5,
                              font=dict(size=14, color='#1a0030', family='Plus Jakarta Sans'), showarrow=False)])
        st.plotly_chart(fig3, use_container_width=True)

with t2:
    st.markdown('<div class="section-title">📊 Store Monthly Trend</div>', unsafe_allow_html=True)

    top5_stores = data_swc['Total Sale'].nlargest(5).index.tolist()
    sel = st.multiselect("Select Stores (to compare)", stores,
                          default=top5_stores, key="swc_sel")
    if sel:
        fig4 = go.Figure()
        bar_colors = ['#7b1fa2','#e91e63','#ff6f00','#00838f','#2e7d32','#1565c0',
                      '#c62828','#4527a0','#00695c','#6a1b9a','#ad1457','#e65100']
        for i, store in enumerate(sel):
            if store in data_swc.index:
                vals = data_swc.loc[store, MONTHS].fillna(0)
                fig4.add_trace(go.Bar(
                    x=MONTH_SHORT, y=vals.values,
                    name=store.replace("SS, ",""),
                    marker=dict(color=bar_colors[i % len(bar_colors)]),
                    hovertemplate='<b>%{fullData.name}</b><br>%{x}: ₹%{y:,.0f}<extra></extra>',
                ))
        fig4.update_layout(**chart_layout(420, "Store-wise Monthly Sale Comparison"),
                          barmode='group', bargap=0.15, bargroupgap=0.05)
        st.plotly_chart(fig4, use_container_width=True)

    st.markdown('<div class="section-title">📋 SWC Data Table</div>', unsafe_allow_html=True)
    swc_disp = data_swc.copy()
    swc_disp.index.name = "Store Name"
    for col in MONTHS:
        swc_disp[col] = swc_disp[col].apply(lambda x: int(x) if pd.notna(x) and x!=0 else "")
    swc_disp['Total Sale'] = swc_disp['Total Sale'].apply(lambda x: int(x) if pd.notna(x) else "")
    swc_disp['Feb Closing Stk'] = swc_disp['Feb Closing Stk'].apply(lambda x: int(x) if pd.notna(x) and x!=0 else "")
    swc_disp['Sale Cont.'] = swc_disp['Sale Cont.'].apply(lambda x: pct(x,4) if pd.notna(x) else "")
    st.dataframe(swc_disp, use_container_width=True)

with t3:
    ca2, cb2 = st.columns([1,1])
    with ca2:
        st.markdown('<div class="section-title">📦 Category Sale vs Stock</div>', unsafe_allow_html=True)
        fig5 = go.Figure()
        cats = list(avail)
        sale_vals = [gt_s[c] for c in cats]
        stk_vals  = [gt_k[c] for c in cats]

        fig5.add_trace(go.Bar(
            name='Sale (MRP)', y=cats, x=sale_vals,
            orientation='h', marker_color='#7b1fa2', marker_line_width=0,
            text=[f"₹{fmt_inr(v)}" for v in sale_vals],
            textposition='outside',
            textfont=dict(size=13, color='#4a0072', family='Inter'),
            hovertemplate='<b>%{y}</b><br>Sale: ₹%{x:,.0f}<extra></extra>'
        ))
        fig5.add_trace(go.Bar(
            name='Closing Stock', y=cats, x=stk_vals,
            orientation='h', marker_color='#ce93d8', marker_line_width=0,
            text=[f"₹{fmt_inr(v)}" for v in stk_vals],
            textposition='outside',
            textfont=dict(size=13, color='#6b21a8', family='Inter'),
            hovertemplate='<b>%{y}</b><br>Stock: ₹%{x:,.0f}<extra></extra>'
        ))
        max_val = max(max(sale_vals), max(stk_vals))
        fig5.update_layout(
            paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
            font=dict(color="#1a0030", family="Inter", size=12),
            margin=dict(l=10, r=160, t=30, b=50),
            legend=dict(font=dict(color="#1a0030", size=11), bgcolor="rgba(255,255,255,0.97)",
                        bordercolor="#ddd6fe", borderwidth=1.5, orientation='h', y=-0.12, x=0.3),
            barmode='group', bargap=0.25, bargroupgap=0.05, height=540,
            title=dict(text="<b>Total Sale vs Closing Stock by Category</b>",
                       font=dict(color='#1a0030', size=15, family='Plus Jakarta Sans')),
            xaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=11),
                       linecolor='#ddd6fe', range=[0, max_val * 1.38]),
            yaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=12),
                       linecolor='#ddd6fe', autorange='reversed'),
        )
        st.plotly_chart(fig5, use_container_width=True)

    with cb2:
        st.markdown('<div class="section-title">🏪 Stores Per Category</div>', unsafe_allow_html=True)
        cat_store_count = (cwc_s[avail] > 0).sum().sort_values(ascending=True)
        fig6 = go.Figure(go.Bar(
            x=cat_store_count.values, y=cat_store_count.index,
            orientation='h',
            marker=dict(color=cat_store_count.values, colorscale=BLUE_SEQ, line=dict(width=0)),
            text=[f"{v} Stores" for v in cat_store_count.values],
            textposition='outside',
            textfont=dict(size=11, color='#1a0030', family='Inter'),
            hovertemplate='<b>%{y}</b><br>%{x} Stores<extra></extra>',
        ))
        fig6.update_layout(paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
            font=dict(color="#1a0030", family="Inter", size=12),
            margin=dict(l=10, r=80, t=50, b=10),
            legend=dict(font=dict(color="#1a0030",size=11), bgcolor="rgba(255,255,255,0.97)"),
            height=420,
            title=dict(text="<b>Store Coverage by Category</b>", font=dict(color='#1a0030', size=15, family='Plus Jakarta Sans')),
            xaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=11), linecolor='#ddd6fe',
                       range=[0, cat_store_count.max() * 1.28]),
            yaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=11), linecolor='#ddd6fe'))
        st.plotly_chart(fig6, use_container_width=True)

    st.markdown('<div class="section-title">📋 CWC Table — Sale</div>', unsafe_allow_html=True)
    cwc_disp = cwc_s.copy()
    cwc_disp.index.name = "Store Name"
    for col in avail + ['TOTAL']:
        cwc_disp[col] = cwc_disp[col].apply(lambda x: int(x) if x!=0 else "")
    st.dataframe(cwc_disp, use_container_width=True)

with t4:
    st.markdown('<div class="section-title">🔥 Store × Category Heatmap</div>', unsafe_allow_html=True)

    hmap = cwc_s[avail].fillna(0).replace(0, np.nan)
    store_labels = [s.replace("SS, ","") for s in hmap.index]

    fig7 = go.Figure(go.Heatmap(
        z=hmap.values,
        x=avail,
        y=store_labels,
        colorscale=[[0,'#fdf8ff'],[0.2,'#e9d8f8'],[0.5,'#c084fc'],[0.75,'#9333ea'],[1,'#581c87']],
        text=[[f"₹{fmt_inr(v)}" if not np.isnan(v) else "" for v in row] for row in hmap.values],
        texttemplate="%{text}", textfont=dict(size=10, color='#1a0030', family='Inter'),
        hoverongaps=False,
        hovertemplate='<b>%{y}</b><br>%{x}<br>₹%{z:,.0f}<extra></extra>',
        showscale=True,
        colorbar=dict(tickfont=dict(color='#9575cd', size=9),
                      outlinecolor='#e8def8', outlinewidth=1,
                      title=dict(text="Sale ₹", font=dict(color='#4a0072', size=10)))
    ))
    fig7.update_layout(paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
        font=dict(color="#1a0030", family="Inter", size=12),
        legend=dict(font=dict(color="#1a0030",size=11), bgcolor="rgba(255,255,255,0.97)"),
        height=700,
        title=dict(text="<b>Sale Heatmap: Store × Category (MRP Value)</b>", font=dict(color='#1a0030', size=15, family='Plus Jakarta Sans')),
        xaxis=dict(tickangle=-30, tickfont=dict(size=11, color='#1a0030', family='Inter'), gridcolor='#ede9fe', linecolor='#ddd6fe'),
        yaxis=dict(tickfont=dict(size=11, color='#1a0030', family='Inter'), autorange='reversed', gridcolor='#ede9fe', linecolor='#ddd6fe'),
        margin=dict(l=130, r=30, t=60, b=80),
    )
    st.plotly_chart(fig7, use_container_width=True)

    hmap_data = cwc_s[avail].fillna(0)
    max_val_hm = hmap_data.max().max()
    max_cat_hm = hmap_data.max().idxmax()
    max_store_hm = hmap_data[max_cat_hm].idxmax()
    zero_store_cats = {}
    for _s in hmap_data.index:
        zero_c = [c for c in avail if hmap_data.loc[_s,c] == 0]
        if len(zero_c) >= 7:
            zero_store_cats[_s.replace("SS, ","")] = len(zero_c)
    cat_totals_hm = hmap_data.sum().sort_values(ascending=False)
    top_cats_hm = cat_totals_hm.head(3)
    store_totals_hm = hmap_data.sum(axis=1).sort_values()
    weakest_store_hm = store_totals_hm.index[0].replace("SS, ","")
    weakest_val_hm = store_totals_hm.values[0]
    best_store_hm = store_totals_hm.index[-1].replace("SS, ","")
    best_val_hm = store_totals_hm.values[-1]
    zero_stores_str = ", ".join([f"<b>{s}</b> ({n} categories zero)" for s,n in list(zero_store_cats.items())[:3]]) if zero_store_cats else "None"
    top_cats_str = "<br>".join([f"<b>{c}</b> — ₹{fmt_inr(int(v))}" for c,v in top_cats_hm.items()])

    st.markdown(f"""
    <div style="background:#f8faff;border:1.5px solid #c7d7f9;border-radius:12px;padding:1.1rem 1.4rem;margin-top:.8rem">
      <div style="font-size:.65rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#1e40af;margin-bottom:.8rem;">
        🔥 HEATMAP — KEY INSIGHTS
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:.8rem;">
        <div style="background:#f5f3ff;border-radius:8px;padding:.8rem 1rem;border-left:4px solid #7c3aed;">
          <div style="font-size:.72rem;font-weight:700;color:#4c1d95;margin-bottom:.3rem;">🏆 HIGHEST SALE COMBINATION</div>
          <div style="font-size:.9rem;font-weight:800;color:#1a0030;"><b>{max_store_hm}</b> → {max_cat_hm}</div>
          <div style="font-size:.82rem;color:#4c1d95;margin-top:.2rem;">₹{fmt_inr(int(max_val_hm))} — Top performing store-category pair</div>
        </div>
        <div style="background:#fef2f2;border-radius:8px;padding:.8rem 1rem;border-left:4px solid #dc2626;">
          <div style="font-size:.72rem;font-weight:700;color:#991b1b;margin-bottom:.3rem;">⚠️ WEAK STORES (Many Zero Categories)</div>
          <div style="font-size:.82rem;color:#1a0030;">{zero_stores_str if zero_store_cats else f"<b>{weakest_store_hm}</b> — Weakest store · ₹{fmt_inr(int(weakest_val_hm))}"}</div>
        </div>
        <div style="background:#f0fdf4;border-radius:8px;padding:.8rem 1rem;border-left:4px solid #16a34a;">
          <div style="font-size:.72rem;font-weight:700;color:#166534;margin-bottom:.3rem;">📦 TOP CATEGORIES BY SALE</div>
          <div style="font-size:.82rem;color:#1a0030;">{top_cats_str}</div>
        </div>
        <div style="background:#fefce8;border-radius:8px;padding:.8rem 1rem;border-left:4px solid #ca8a04;">
          <div style="font-size:.72rem;font-weight:700;color:#854d0e;margin-bottom:.3rem;">💡 HOW TO READ THIS HEATMAP</div>
          <div style="font-size:.82rem;color:#1a0030;">
            <b>Dark Purple</b> = High sale ✅<br>
            <b>Light Pink</b> = Low sale ⚠️<br>
            <b>Blank/—</b> = Zero sale ❌<br>
            Focus dark cells = your best opportunities
          </div>
        </div>
      </div>
      <div style="margin-top:.8rem;padding:.6rem .8rem;background:#fff;border-radius:8px;font-size:.78rem;color:#374151;">
        📌 <b>Best Store:</b> {best_store_hm} (₹{fmt_inr(int(best_val_hm))}) &nbsp;·&nbsp;
        <b>Weakest Store:</b> {weakest_store_hm} (₹{fmt_inr(int(weakest_val_hm))}) &nbsp;·&nbsp;
        <b>Top Category:</b> {cat_totals_hm.index[0]} (₹{fmt_inr(int(cat_totals_hm.values[0]))})
      </div>
    </div>
    """, unsafe_allow_html=True)

with t5:
    st.markdown('<div class="section-title">🔍 Store Deep Dive</div>', unsafe_allow_html=True)

    col_sel, col_rank = st.columns([1, 1])
    with col_sel:
        sel_store = st.selectbox("Select Store", stores, key="dd_store")

    if sel_store and sel_store in data_swc.index:
        row = data_swc.loc[sel_store]
        total_s = row['Total Sale']
        closing_s = row['Feb Closing Stk']
        cont_s_val = row['Sale Cont.']
        rank = int(data_swc['Total Sale'].rank(ascending=False)[sel_store])

        m1, m2, m3, m4 = st.columns(4)
        for col, lbl, val, icon, sub in [
            (m1, "Total MRP Sale",    f"₹{fmt_inr(total_s)}",   "💰", f"Apr'25 → Feb'26 Total"),
            (m2, "Feb Closing Stock", f"₹{fmt_inr(closing_s)}", "📦", f"Feb 2026 · <span style='font-size:.95rem;font-weight:800;color:#fff'>{int(stock[(stock['Store Name']==sel_store) & (stock['Month']=='Feb Closing')]['Quantity'].sum()) if 'Quantity' in stock.columns else ''} Pcs</span>"),
            (m3, "Sale Contribution", pct(cont_s_val,4),        "📊", "Of Total Sale"),
        ]:
            with col:
                st.markdown(f"""<div class="kpi-card">
                  <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:.2rem">
                    <div class="kpi-label">{lbl}</div>
                    <span style="font-size:1.3rem;opacity:0.85">{icon}</span>
                  </div>
                  <div class="kpi-value blue">{val}</div>
                  <div class="kpi-sub">{sub}</div>
                </div>""", unsafe_allow_html=True)

        with m4:
            st.markdown(f"""<div class="kpi-card">
              <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:.2rem">
                <div class="kpi-label">Store Rank</div>
                <span style="font-size:1.3rem;opacity:0.85">🏅</span>
              </div>
              <div class="kpi-value blue">#{rank}</div>
              <div class="kpi-sub">{sel_store.replace("SS, ","SS ")} &nbsp;·&nbsp; Out of {len(stores)}</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        da, db = st.columns([3,2])

        with da:
            monthly_vals = row[MONTHS].fillna(0)
            fig8 = go.Figure(go.Bar(
                x=MONTH_SHORT, y=monthly_vals.values,
                marker=dict(color=monthly_vals.values, colorscale=BLUE_SEQ, line=dict(width=0)),
                text=[f"₹{fmt_inr(v)}" if v > 0 else "" for v in monthly_vals.values],
                textposition='outside', textfont=dict(size=10, color='#1a0030', family='Inter'),
                hovertemplate='<b>%{x}</b><br>₹%{y:,.0f}<extra></extra>',
            ))
            fig8.update_layout(**chart_layout(300, f"{sel_store} — Monthly Sale"), bargap=0.3)
            st.plotly_chart(fig8, use_container_width=True)

        with db:
            if sel_store in cwc_s.index:
                cat_vals2 = cwc_s.loc[sel_store, avail]
                cat_vals2 = cat_vals2[cat_vals2 > 0]
                if len(cat_vals2) > 0:
                    fig9 = go.Figure(go.Pie(
                        labels=cat_vals2.index, values=cat_vals2.values, hole=0.48,
                        marker=dict(colors=CAT_COLORS_LIGHT[:len(cat_vals2)],
                                    line=dict(color='#ffffff', width=3)),
                        textinfo='label+percent',
                        textfont=dict(size=11, color='#1a0030', family='Inter'),
                        insidetextfont=dict(size=10, color='#ffffff'),
                        textposition='auto',
                        hovertemplate='<b>%{label}</b><br>₹%{value:,.0f}<extra></extra>',
                    ))
                    fig9.update_layout(**chart_layout(300, "Category Mix"))
                    st.plotly_chart(fig9, use_container_width=True)

with t6:
    st.markdown('<div class="section-title">📊 Month-on-Month Growth — All Stores</div>', unsafe_allow_html=True)

    mom_data = data_swc[MONTHS].copy()
    mom_pct = mom_data.pct_change(axis=1) * 100
    monthly_total = data_swc[MONTHS].sum()
    mom_total = monthly_total.pct_change() * 100

    apr_sale = float(monthly_total.values[0])
    mom_vals_plot = [float(v) for v in mom_total.values[1:]]
    mom_labels_plot = MONTH_SHORT[1:]
    colors_mom = ['#16a34a' if v >= 0 else '#dc2626' for v in mom_vals_plot]
    mom_texts = [f"{v:+.1f}%" for v in mom_vals_plot]

    fig_mom = go.Figure(go.Bar(
        x=mom_labels_plot,
        y=mom_vals_plot,
        marker=dict(color=colors_mom, line=dict(width=0)),
        text=mom_texts,
        textposition='outside',
        textfont=dict(size=13, color='#1a0030', family='Inter'),
        hovertemplate='<b>%{x}</b><br>Growth: %{y:+.1f}%<extra></extra>',
    ))
    fig_mom.update_layout(
        paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
        font=dict(color="#1a0030", family="Inter", size=12),
        height=360, bargap=0.3,
        margin=dict(l=20, r=20, t=50, b=40),
        title=dict(text="<b>Month-on-Month Sale Growth (%) — All Stores Combined</b>",
                   font=dict(color='#1a0030', size=15, family='Plus Jakarta Sans')),
        xaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=11), linecolor='#ddd6fe'),
        yaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=11), linecolor='#ddd6fe',
                   zeroline=True, zerolinecolor='#9c27b0', zerolinewidth=2),
    )
    st.markdown(f"""<div style="background:linear-gradient(135deg,#f5f3ff,#ede9fe);border:1.5px solid #9c27b0;
        border-radius:10px;padding:.6rem 1.2rem;margin-bottom:.6rem;display:inline-block">
        <span style="font-size:.65rem;font-weight:700;letter-spacing:2px;color:#6b21a8;text-transform:uppercase">
        📌 APR'25 — START MONTH</span>
        <span style="font-size:1rem;font-weight:800;color:#4c1d95;margin-left:1rem">₹{fmt_inr(int(apr_sale))}</span>
        <span style="font-size:.75rem;color:#7c3aed;margin-left:.5rem">(Base — no previous month to compare)</span>
    </div>""", unsafe_allow_html=True)
    st.plotly_chart(fig_mom, use_container_width=True)

    ca6, cb6 = st.columns(2)
    with ca6:
        st.markdown('<div class="section-title">🏆 Top 5 Stores</div>', unsafe_allow_html=True)
        top5 = data_swc['Total Sale'].nlargest(5).sort_values()
        fig_top5 = go.Figure(go.Bar(
            x=top5.values, y=top5.index.str.replace("SS, ", ""),
            orientation='h',
            marker=dict(color='#16a34a', line=dict(width=0)),
            text=[f"₹{fmt_inr(v)}" for v in top5.values],
            textposition='outside',
            textfont=dict(size=12, color='#1a0030', family='Inter'),
            hovertemplate='<b>%{y}</b><br>₹%{x:,.0f}<extra></extra>',
        ))
        fig_top5.update_layout(
            paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
            font=dict(color="#1a0030", family="Inter", size=12),
            height=320, margin=dict(l=10, r=160, t=40, b=20),
            title=dict(text="<b>Top 5 Stores by Total Sale</b>",
                       font=dict(color='#1a0030', size=14, family='Plus Jakarta Sans')),
            xaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=10),
                       linecolor='#ddd6fe', range=[0, top5.max() * 1.4]),
            yaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=11), linecolor='#ddd6fe'),
        )
        st.plotly_chart(fig_top5, use_container_width=True)

    with cb6:
        st.markdown('<div class="section-title">🔴 Bottom 5 Stores</div>', unsafe_allow_html=True)
        bot5 = data_swc['Total Sale'].nsmallest(5).sort_values(ascending=False)
        fig_bot5 = go.Figure(go.Bar(
            x=bot5.values, y=bot5.index.str.replace("SS, ", ""),
            orientation='h',
            marker=dict(color='#dc2626', line=dict(width=0)),
            text=[f"₹{fmt_inr(v)}" for v in bot5.values],
            textposition='outside',
            textfont=dict(size=12, color='#1a0030', family='Inter'),
            hovertemplate='<b>%{y}</b><br>₹%{x:,.0f}<extra></extra>',
        ))
        fig_bot5.update_layout(
            paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
            font=dict(color="#1a0030", family="Inter", size=12),
            height=320, margin=dict(l=10, r=160, t=40, b=20),
            title=dict(text="<b>Bottom 5 Stores by Total Sale</b>",
                       font=dict(color='#1a0030', size=14, family='Plus Jakarta Sans')),
            xaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=10),
                       linecolor='#ddd6fe', range=[0, bot5.max() * 1.5]),
            yaxis=dict(gridcolor='#ede9fe', tickfont=dict(color='#1a0030', size=11), linecolor='#ddd6fe'),
        )
        st.plotly_chart(fig_bot5, use_container_width=True)

    st.markdown('<div class="section-title">❌ Zero Sale Months — Store-wise</div>', unsafe_allow_html=True)
    zero_records = []
    for store in stores:
        row = data_swc.loc[store, MONTHS]
        zero_months = [MONTH_SHORT[i] for i, v in enumerate(row.values) if pd.isna(v) or v == 0]
        sale_months_st = [MONTH_SHORT[i] for i, v in enumerate(row.values) if not (pd.isna(v) or v == 0)]
        if zero_months:
            zero_records.append({
                'Store': store.replace("SS, ", ""),
                'Sale Months': ', '.join(sale_months_st) if sale_months_st else '—',
                'Sale Count': len(sale_months_st),
                'Zero Sale Months': ', '.join(zero_months),
                'Zero Count': len(zero_months),
                'Total Sale': f"₹{fmt_inr(data_swc.loc[store, 'Total Sale'])}"
            })
    if zero_records:
        zero_df = pd.DataFrame(zero_records).sort_values('Zero Count', ascending=False)
        st.dataframe(zero_df, use_container_width=True, hide_index=True)
        st.markdown(f"""<div style="background:#fef2f2;border:1px solid #fca5a5;border-radius:10px;
            padding:.6rem 1rem;font-size:.8rem;color:#991b1b;margin-top:.5rem">
            ⚠️ <b>{len(zero_records)} stores</b> had zero sale in one or more months.
        </div>""", unsafe_allow_html=True)

    st.markdown('<div class="section-title">📈 Month-on-Month Growth — Store-wise (Last Month vs Previous)</div>', unsafe_allow_html=True)
    mom_store = []
    for store in stores:
        row = data_swc.loc[store, MONTHS]
        last = row.values[-1]; prev = row.values[-2]
        if prev > 0:
            growth = ((last - prev) / prev) * 100
            arrow = "🟢 ▲" if growth >= 0 else "🔴 ▼"
        else:
            growth = None; arrow = "➖ N/A"
        mom_store.append({
            'Store': store.replace("SS, ", ""),
            'Jan Sale': f"₹{fmt_inr(prev)}",
            'Feb Sale': f"₹{fmt_inr(last)}",
            'MoM Growth': f"{growth:+.1f}%" if growth is not None else "N/A",
            'Trend': arrow
        })
    mom_df = pd.DataFrame(mom_store)
    st.dataframe(mom_df, use_container_width=True, hide_index=True)


with t7:
    st.markdown('<div class="section-title">📊 Sell-Through Rate — Store × Category</div>', unsafe_allow_html=True)
    st.markdown("""<div style="background:#f0fdf4;border:1px solid #86efac;border-radius:8px;
        padding:.5rem 1rem;font-size:.8rem;color:#166534;margin-bottom:.8rem">
        <b>Sell-Through Rate</b> = Sale ÷ (Sale + Closing Stock) × 100 &nbsp;·&nbsp;
        Higher % = Better selling &nbsp;·&nbsp; Lower % = Overstocked
    </div>""", unsafe_allow_html=True)

    st_matrix = pd.DataFrame(index=cwc_s.index, columns=avail)
    for store in cwc_s.index:
        for cat in avail:
            s = cwc_s.loc[store, cat] if store in cwc_s.index else 0
            k = cwc_k.loc[store, cat] if store in cwc_k.index else 0
            total = s + k
            st_matrix.loc[store, cat] = round(s / total * 100, 1) if total > 0 else 0
    st_matrix = st_matrix.astype(float)

    fig_st = go.Figure(go.Heatmap(
        z=st_matrix.values.tolist(),
        x=list(avail),
        y=st_matrix.index.str.replace("SS, ", "").tolist(),
        colorscale=[[0, '#fef2f2'], [0.3, '#fca5a5'], [0.6, '#fde68a'], [0.8, '#86efac'], [1, '#16a34a']],
        text=[[f"{v:.0f}%" if v > 0 else "—" for v in row] for row in st_matrix.values.tolist()],
        texttemplate="%{text}",
        textfont=dict(size=9, color='#1a0030'),
        hoverongaps=False,
        hovertemplate='<b>%{y}</b><br>%{x}<br>Sell-Through: %{z:.1f}%<extra></extra>',
        zmin=0, zmax=100,
        colorbar=dict(title="ST%", tickfont=dict(color='#1a0030', size=9), ticksuffix="%")
    ))
    fig_st.update_layout(
        paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(255,255,255,1)",
        font=dict(color="#1a0030", family="Inter", size=11),
        height=700, margin=dict(l=150, r=30, t=50, b=80),
        title=dict(text="<b>Sell-Through Rate (%) — Store × Category</b>",
                   font=dict(color='#1a0030', size=15, family='Plus Jakarta Sans')),
        xaxis=dict(tickangle=-30, tickfont=dict(size=10, color='#1a0030')),
        yaxis=dict(tickfont=dict(size=10, color='#1a0030'), autorange='reversed'),
    )
    st.plotly_chart(fig_st, use_container_width=True)

    avg_st = st_matrix.replace(0, np.nan).stack().mean()
    store_avg_st = st_matrix.replace(0, np.nan).mean(axis=1).sort_values(ascending=False)
    best_stores = store_avg_st[store_avg_st >= 60].index.str.replace("SS, ", "").tolist()
    worst_stores = store_avg_st[store_avg_st < 30].index.str.replace("SS, ", "").tolist()
    cat_avg_st = st_matrix.replace(0, np.nan).mean(axis=0).sort_values(ascending=False)
    best_cats = cat_avg_st[cat_avg_st >= 60].index.tolist()
    slow_cats = cat_avg_st[cat_avg_st < 30].index.tolist()

    perfect_stores = []
    for store in st_matrix.index:
        vals = st_matrix.loc[store].replace(0, np.nan).dropna()
        if len(vals) > 0 and vals.min() >= 95:
            perfect_stores.append(store.replace("SS, ", ""))

    restock_needed = []
    for store in cwc_s.index:
        for cat in avail:
            s = cwc_s.loc[store, cat]
            k = cwc_k.loc[store, cat] if store in cwc_k.index else 0
            total = s + k
            if total > 0 and s > 5000:
                st_r = s / total * 100
                if st_r >= 80:
                    restock_needed.append(f"{store.replace('SS, ','')} → {cat}")

    top_heroes = (perfect_stores or best_stores[:3])
    hero_lines = "<br>".join([f"<b>{s}</b> — Strong sell-through across categories" for s in top_heroes]) if top_heroes else "No store with consistently high ST rate"
    warn_lines = "<br>".join([f"<b>{s}</b> — Low sell-through, overstocked" for s in worst_stores[:3]]) if worst_stores else "All stores performing reasonably well"
    restock_lines = "<br>".join(restock_needed[:4]) if restock_needed else "No urgent restock required"
    restock_cats_st = set([x.split(' → ')[1] for x in restock_needed])
    truly_slow_cats = [c for c in slow_cats if c not in restock_cats_st]
    mixed_cats = [c for c in slow_cats if c in restock_cats_st]
    slow_lines_parts = [f"<b>{c}</b> — Reduce stock (slow across all stores)" for c in truly_slow_cats[:3]]
    if mixed_cats:
        slow_lines_parts.append(f"<i style='color:#ca8a04'>{', '.join(mixed_cats[:2])} — Mixed: slow overall but fast in specific stores</i>")
    slow_lines = "<br>".join(slow_lines_parts) if slow_lines_parts else "No universally slow categories"

    st.markdown(f"""
    <div style="background:#f8faff;border:1.5px solid #c7d7f9;border-radius:12px;padding:1.1rem 1.4rem;margin-bottom:1.2rem;">
      <div style="font-size:.65rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#1e40af;margin-bottom:.8rem;">
        📊 SELL-THROUGH RATE — KEY INSIGHTS
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:.8rem;">
        <div style="background:#f0fdf4;border-radius:8px;padding:.8rem 1rem;border-left:4px solid #16a34a;">
          <div style="font-size:.72rem;font-weight:700;color:#166534;margin-bottom:.3rem;">🏆 BEST PERFORMING STORES</div>
          <div style="font-size:.82rem;color:#1a0030;">{hero_lines}</div>
        </div>
        <div style="background:#fef2f2;border-radius:8px;padding:.8rem 1rem;border-left:4px solid #dc2626;">
          <div style="font-size:.72rem;font-weight:700;color:#991b1b;margin-bottom:.3rem;">⚠️ STORES NEEDING ATTENTION</div>
          <div style="font-size:.82rem;color:#1a0030;">{warn_lines}</div>
        </div>
        <div style="background:#eff6ff;border-radius:8px;padding:.8rem 1rem;border-left:4px solid #1d4ed8;">
          <div style="font-size:.72rem;font-weight:700;color:#1e40af;margin-bottom:.3rem;">📦 URGENT RESTOCK NEEDED</div>
          <div style="font-size:.82rem;color:#1a0030;">{restock_lines}</div>
          <div style="font-size:.7rem;color:#6b7280;margin-top:.3rem;">ST Rate ≥ 80% with significant sale volume</div>
        </div>
        <div style="background:#fefce8;border-radius:8px;padding:.8rem 1rem;border-left:4px solid #ca8a04;">
          <div style="font-size:.72rem;font-weight:700;color:#854d0e;margin-bottom:.3rem;">🐌 SLOW MOVING CATEGORIES</div>
          <div style="font-size:.82rem;color:#1a0030;">{slow_lines}</div>
        </div>
      </div>
      <div style="margin-top:.8rem;padding:.6rem .8rem;background:#fff;border-radius:8px;font-size:.78rem;color:#374151;">
        📌 <b>Overall Avg Sell-Through: {avg_st:.1f}%</b> &nbsp;·&nbsp;
        <span style="color:#16a34a;font-weight:600;">Green ≥ 60%</span> = Good &nbsp;·&nbsp;
        <span style="color:#ca8a04;font-weight:600;">Yellow 30–59%</span> = Average &nbsp;·&nbsp;
        <span style="color:#dc2626;font-weight:600;">Red &lt; 30%</span> = Slow mover
      </div>
    </div>
    """, unsafe_allow_html=True)

    cc7, cd7 = st.columns(2)
    with cc7:
        st.markdown('<div class="section-title">🟢 Category-Store Fit Score</div>', unsafe_allow_html=True)
        st.markdown("""<div style="font-size:.75rem;color:#607d9b;margin-bottom:.6rem">
            Based on Sell-Through Rate: 🟢 Strong (≥60%) &nbsp;·&nbsp; 🟡 Average (30-59%) &nbsp;·&nbsp; 🔴 Weak (&lt;30%)
        </div>""", unsafe_allow_html=True)

        fit_records = []
        for store in cwc_s.index:
            for cat in avail:
                s = cwc_s.loc[store, cat] if store in cwc_s.index else 0
                k = cwc_k.loc[store, cat] if store in cwc_k.index else 0
                total = s + k
                if total > 0:
                    st_rate = s / total * 100
                    fit = "🟢 Strong" if st_rate >= 60 else ("🟡 Average" if st_rate >= 30 else "🔴 Weak")
                    sale_qty = int(sale[(sale['Store Name']==store) & (sale['CATEGORY']==cat)]['Quantity'].sum()) if 'Quantity' in sale.columns else '—'
                    stk_qty = int(stock[(stock['Store Name']==store) & (stock['CATEGORY']==cat) & (stock['Month']=='Feb Closing')]['Quantity'].sum()) if 'Quantity' in stock.columns else '—'
                    fit_records.append({
                        'Store': store.replace("SS, ", ""),
                        'Category': cat,
                        'Sale (₹)': f"₹{fmt_inr(s)}",
                        'Sale Qty': sale_qty,
                        'Stock (₹)': f"₹{fmt_inr(k)}",
                        'Stock Qty': stk_qty,
                        'Sell-Through': f"{st_rate:.1f}%",
                        'Fit': fit
                    })

        fit_df = pd.DataFrame(fit_records)
        fit_filter = st.selectbox("Filter by Fit", ["All", "🟢 Strong", "🟡 Average", "🔴 Weak"], key="fit_filter")
        if fit_filter != "All":
            fit_df = fit_df[fit_df['Fit'] == fit_filter]
        st.dataframe(fit_df, use_container_width=True, hide_index=True, height=400)

    with cd7:
        st.markdown('<div class="section-title">⚠️ Dead Stock Alert</div>', unsafe_allow_html=True)
        st.markdown("""<div style="background:#fef2f2;border:1px solid #fca5a5;border-radius:8px;
            padding:.5rem 1rem;font-size:.8rem;color:#991b1b;margin-bottom:.6rem">
            Dead Stock = Stock exists but ZERO sale in entire period. Immediate action needed!
        </div>""", unsafe_allow_html=True)

        dead_records = []
        for store in cwc_s.index:
            for cat in avail:
                s = cwc_s.loc[store, cat] if store in cwc_s.index else 0
                k = cwc_k.loc[store, cat] if store in cwc_k.index else 0
                if s == 0 and k > 0:
                    dq = int(stock[(stock['Store Name']==store) & (stock['CATEGORY']==cat) & (stock['Month']=='Feb Closing')]['Quantity'].sum()) if 'Quantity' in stock.columns else '—'
                    dead_records.append({
                        'Store': store.replace("SS, ", ""),
                        'Category': cat,
                        'Dead Stock (₹)': f"₹{fmt_inr(k)}",
                        'Dead Stock Qty': f"{dq} Pcs",
                        'Action': '⚠️ Transfer or Liquidate'
                    })

        if dead_records:
            dead_df = pd.DataFrame(dead_records).sort_values('Store')
            st.dataframe(dead_df, use_container_width=True, hide_index=True, height=400)
            total_dead_val = 0
            total_dead_qty = 0
            for r in dead_records:
                sname = "SS, " + r['Store'] if "SS, " + r['Store'] in cwc_k.index else r['Store']
                if sname in cwc_k.index:
                    total_dead_val += cwc_k.loc[sname, r['Category']]
                if 'Quantity' in stock.columns:
                    sname_full = "SS, " + r['Store']
                    q = stock[(stock['Store Name']==sname_full) & (stock['CATEGORY']==r['Category']) & (stock['Month']=='Feb Closing')]['Quantity'].sum()
                    total_dead_qty += int(q)
            st.markdown(f"""<div style="background:#fef2f2;border:1px solid #fca5a5;border-radius:10px;
                padding:.8rem 1.1rem;margin-top:.5rem">
                <div style="font-size:.85rem;font-weight:700;color:#991b1b;margin-bottom:.4rem">
                    ⚠️ <b>{len(dead_records)} store-category combinations</b> have Dead Stock
                </div>
                <div style="display:flex;gap:2rem;font-size:.82rem;color:#7f1d1d">
                    <span>💰 Total Dead Stock Value: <b>₹{fmt_inr(int(total_dead_val))}</b></span>
                    <span>📦 Total Dead Qty: <b>{total_dead_qty} Pcs</b></span>
                </div>
            </div>""", unsafe_allow_html=True)
        else:
            st.success("✅ No dead stock found!")

    st.markdown('<div class="section-title">📋 Stock Recommendations</div>', unsafe_allow_html=True)
    rec_records = []
    for store in cwc_s.index:
        for cat in avail:
            s = cwc_s.loc[store, cat] if store in cwc_s.index else 0
            k = cwc_k.loc[store, cat] if store in cwc_k.index else 0
            total = s + k
            if total == 0: continue
            st_rate = s / total * 100
            if st_rate >= 75 and s > 5000:
                rec = "📦 Increase Stock — High Demand"; priority = "🔴 Urgent"
            elif st_rate >= 60 and s > 0:
                rec = "📦 Replenish Stock — Good Seller"; priority = "🟡 Medium"
            elif st_rate < 20 and k > 10000:
                rec = "🔄 Transfer to Better Performing Store"; priority = "🔴 Urgent"
            elif s == 0 and k > 0:
                rec = "❌ Remove Stock — No Sale"; priority = "🔴 Urgent"
            elif st_rate < 30 and k > 5000:
                rec = "⬇️ Reduce Stock — Slow Mover"; priority = "🟡 Medium"
            else: continue
            rsq = int(sale[sale['Store Name']==store][sale['CATEGORY']==cat]['Quantity'].sum()) if 'Quantity' in sale.columns else '—'
            rkq = int(stock[(stock['Store Name']==store) & (stock['CATEGORY']==cat) & (stock['Month']=='Feb Closing')]['Quantity'].sum()) if 'Quantity' in stock.columns else '—'
            rec_records.append({
                'Store': store.replace("SS, ", ""),
                'Category': cat,
                'Sale (₹)': f"₹{fmt_inr(s)}",
                'Sale Qty': rsq,
                'Stock (₹)': f"₹{fmt_inr(k)}",
                'Stock Qty': rkq,
                'Sell-Through': f"{st_rate:.1f}%",
                'Recommendation': rec,
                'Priority': priority
            })

    if rec_records:
        rec_df = pd.DataFrame(rec_records).sort_values(['Priority', 'Store'])
        priority_filter = st.selectbox("Filter by Priority", ["All", "🔴 Urgent", "🟡 Medium"], key="rec_priority")
        if priority_filter != "All":
            rec_df = rec_df[rec_df['Priority'] == priority_filter]
        st.dataframe(rec_df, use_container_width=True, hide_index=True)
        st.markdown(f"""<div style="background:#eff6ff;border:1px solid #93c5fd;border-radius:8px;
            padding:.6rem 1rem;font-size:.8rem;color:#1e40af;margin-top:.5rem">
            💡 <b>{len(rec_records)} recommendations</b> generated based on Sell-Through Rate analysis.
        </div>""", unsafe_allow_html=True)

with t8:
    st.markdown('<div class="section-title">🤖 AI Strategy Summary</div>', unsafe_allow_html=True)
    st.markdown("""<div style="background:linear-gradient(135deg,#3a0068,#6a1b9a);border-radius:12px;
        padding:1rem 1.4rem;margin-bottom:1rem;color:#fff">
        <div style="font-size:.65rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;
            color:rgba(255,255,255,.7);margin-bottom:.3rem">HOW IT WORKS</div>
        <div style="font-size:.85rem">AI analyses your complete sales data — stores, categories,
        sell-through rates, dead stock — and generates a smart business strategy with action plan.</div>
    </div>""", unsafe_allow_html=True)

    st_mat = pd.DataFrame(index=cwc_s.index, columns=avail)
    for _s in cwc_s.index:
        for _c in avail:
            sv = float(cwc_s.loc[_s, _c])
            kv = float(cwc_k.loc[_s, _c]) if _s in cwc_k.index else 0
            t = sv + kv
            st_mat.loc[_s, _c] = round(sv/t*100, 1) if t > 0 else 0
    st_mat = st_mat.astype(float)

    total_sale = float(grand)
    total_stock = float(data_swc['Feb Closing Stk'].sum())
    overall_st = float(st_mat.replace(0, np.nan).stack().mean())
    num_stores = len(stores)
    num_cats = len(avail)

    top3 = data_swc['Total Sale'].nlargest(3)
    bot3 = data_swc['Total Sale'].nsmallest(3)
    cat_sale = gt_s[avail].sort_values(ascending=False)
    cat_stock = gt_k[avail].sort_values(ascending=False)
    cat_st = st_mat.replace(0, np.nan).mean(axis=0).sort_values(ascending=False)

    dead_items = []
    dead_val = 0
    for _s in cwc_s.index:
        for _c in avail:
            sv = float(cwc_s.loc[_s, _c])
            kv = float(cwc_k.loc[_s, _c]) if _s in cwc_k.index else 0
            if sv == 0 and kv > 0:
                dead_items.append(f"{_s.replace('SS, ','')}:{_c}=₹{fmt_inr(kv)}")
                dead_val += kv

    restock_items = []
    for _s in cwc_s.index:
        for _c in avail:
            sv = float(cwc_s.loc[_s, _c])
            kv = float(cwc_k.loc[_s, _c]) if _s in cwc_k.index else 0
            t = sv + kv
            if t > 0 and sv > 5000 and sv/t*100 >= 80:
                restock_items.append(f"{_s.replace('SS, ','')}:{_c}({sv/t*100:.0f}%ST)")

    monthly_vals = data_swc[MONTHS].sum()
    mom_last = ((monthly_vals.values[-1] - monthly_vals.values[-2]) / monthly_vals.values[-2] * 100) if monthly_vals.values[-2] > 0 else 0
    best_month_idx = monthly_vals.values.argmax()
    worst_month_idx = monthly_vals.values.argmin()

    zero_stores = []
    for _s in stores:
        row = data_swc.loc[_s, MONTHS]
        zc = sum(1 for v in row.values if pd.isna(v) or v == 0)
        if zc >= 3:
            zero_stores.append(f"{_s.replace('SS, ','')}({zc} months)")

    data_prompt = f"""You are a senior retail business consultant analyzing sales data for SS Retail (UCB brand stores in India).

BUSINESS DATA SUMMARY:
- Total MRP Sale: ₹{fmt_inr(int(total_sale))} across {num_stores} stores, {num_cats} categories, Apr'25–Feb'26 (11 months)
- Total Closing Stock: ₹{fmt_inr(int(total_stock))}
- Overall Sell-Through Rate: {overall_st:.1f}%
- Last Month Growth (Jan→Feb): {mom_last:+.1f}%
- Best Month: {MONTH_SHORT[best_month_idx]} | Worst Month: {MONTH_SHORT[worst_month_idx]}

TOP 3 STORES (by sale):
{chr(10).join([f"- {s.replace('SS, ','')}: ₹{fmt_inr(int(v))} ({v/total_sale*100:.1f}% of total)" for s,v in top3.items()])}

BOTTOM 3 STORES (by sale):
{chr(10).join([f"- {s.replace('SS, ','')}: ₹{fmt_inr(int(v))} ({v/total_sale*100:.1f}% of total)" for s,v in bot3.items()])}

TOP 5 CATEGORIES (by sale):
{chr(10).join([f"- {c}: ₹{fmt_inr(int(cat_sale[c]))} | ST Rate: {cat_st.get(c,0):.1f}%" for c in cat_sale.index[:5]])}

BOTTOM 3 CATEGORIES (by sale):
{chr(10).join([f"- {c}: ₹{fmt_inr(int(cat_sale[c]))} | ST Rate: {cat_st.get(c,0):.1f}%" for c in cat_sale.index[-3:]])}

DEAD STOCK (stock exists, zero sale):
Total Dead Stock Value: ₹{fmt_inr(int(dead_val))}
Items: {', '.join(dead_items[:8]) if dead_items else 'None'}

URGENT RESTOCK NEEDED (ST Rate ≥80%):
{', '.join(restock_items[:8]) if restock_items else 'None'}

STORES WITH POOR CONSISTENCY (3+ months zero sale):
{', '.join(zero_stores[:6]) if zero_stores else 'None'}

Based on this data, provide a structured business strategy report in English with these exact sections:

1. EXECUTIVE SUMMARY (2-3 sentences — overall business health)
2. KEY STRENGTHS (3 bullet points — what is working well)
3. CRITICAL ISSUES (3 bullet points — biggest problems)
4. HOW TO INCREASE SALE — STRATEGIES (5 specific, actionable strategies with expected impact)
5. INVENTORY ACTION PLAN (specific store/category actions — restock, transfer, liquidate)
6. IMMEDIATE PRIORITIES (3 actions for THIS WEEK, 3 for THIS MONTH, 3 for NEXT QUARTER)

Be specific with store names and numbers. Be direct and business-focused. No fluff."""

    if 'ai_summary' not in st.session_state:
        st.session_state.ai_summary = None
    if 'ai_loading' not in st.session_state:
        st.session_state.ai_loading = False

    col_btn1, col_btn2, col_btn3 = st.columns([1,2,1])
    with col_btn2:
        if st.button("🤖  Generate AI Strategy Summary", use_container_width=True):
            st.session_state.ai_loading = True
            st.session_state.ai_summary = None

    if st.session_state.ai_loading and st.session_state.ai_summary is None:
        with st.spinner("🤖 AI is analysing your data... please wait..."):
            try:
                import requests
                response = requests.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"Content-Type": "application/json"},
                    json={
                        "model": "claude-sonnet-4-20250514",
                        "max_tokens": 1500,
                        "messages": [{"role": "user", "content": data_prompt}]
                    },
                    timeout=60
                )
                if response.status_code == 200:
                    result = response.json()
                    ai_text = result['content'][0]['text']
                    st.session_state.ai_summary = ai_text
                    st.session_state.ai_loading = False
                else:
                    st.session_state.ai_summary = f"API Error: {response.status_code} — {response.text[:200]}"
                    st.session_state.ai_loading = False
            except Exception as e:
                st.session_state.ai_summary = f"Error: {str(e)}"
                st.session_state.ai_loading = False
        st.rerun()

    if st.session_state.ai_summary:
        summary_text = st.session_state.ai_summary
        section_styles = {
            "EXECUTIVE SUMMARY": ("📋", "#1e3a5f", "#eff6ff", "#1e40af"),
            "KEY STRENGTHS": ("💪", "#166534", "#f0fdf4", "#16a34a"),
            "CRITICAL ISSUES": ("🚨", "#991b1b", "#fef2f2", "#dc2626"),
            "HOW TO INCREASE SALE": ("🚀", "#4c1d95", "#f5f3ff", "#7c3aed"),
            "INVENTORY ACTION PLAN": ("📦", "#854d0e", "#fefce8", "#ca8a04"),
            "IMMEDIATE PRIORITIES": ("⚡", "#065f46", "#ecfdf5", "#059669"),
        }
        lines = summary_text.split('\n')
        current_section = None
        sections = {}
        current_content = []
        for line in lines:
            line = line.strip()
            if not line: continue
            matched = False
            for sec_key in section_styles.keys():
                if sec_key in line.upper():
                    if current_section:
                        sections[current_section] = '\n'.join(current_content)
                    current_section = sec_key
                    current_content = []
                    matched = True
                    break
            if not matched and current_section:
                current_content.append(line)
        if current_section:
            sections[current_section] = '\n'.join(current_content)

        for sec_key, (icon, title_color, bg_color, border_color) in section_styles.items():
            if sec_key in sections:
                sec_content = sections[sec_key]
                formatted = []
                for l in sec_content.split('\n'):
                    l = l.strip()
                    if not l: continue
                    if l.startswith(('-', '•', '*')):
                        l = '• ' + l.lstrip('-•* ').strip()
                    elif l[0].isdigit() and '.' in l[:3]:
                        l = '→ ' + l[2:].strip() if l[1] == '.' else l
                    formatted.append(f'<div style="margin:.25rem 0;font-size:.85rem;color:#1a0030;line-height:1.5">{l}</div>')
                st.markdown(f"""
                <div style="background:{bg_color};border-left:4px solid {border_color};
                    border-radius:10px;padding:.9rem 1.1rem;margin-bottom:.8rem">
                    <div style="font-size:.6rem;font-weight:800;letter-spacing:2px;
                        text-transform:uppercase;color:{title_color};margin-bottom:.5rem">
                        {icon} {sec_key}
                    </div>
                    {''.join(formatted)}
                </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        d1, d2, d3 = st.columns([1,2,1])
        with d2:
            summary_export = f"SS RETAIL — AI STRATEGY SUMMARY\nGenerated: {pd.Timestamp.now().strftime('%d %b %Y %I:%M %p')}\n\n{summary_text}"
            st.download_button(
                "📥  Download Strategy Summary (TXT)",
                data=summary_export,
                file_name="SS_AI_Strategy_Summary.txt",
                mime="text/plain",
                use_container_width=True
            )

    elif not st.session_state.ai_loading:
        st.markdown("""<div style="text-align:center;padding:3rem 0">
            <div style="font-size:3rem">🤖</div>
            <div style="font-size:1rem;color:#607d9b;font-weight:500;margin-top:.8rem">
                Click "Generate AI Strategy Summary" to get AI analysis</div>
            <div style="font-size:.8rem;color:#90a4c0;margin-top:.4rem">
                Analyses all stores · categories · months of data</div>
        </div>""", unsafe_allow_html=True)
